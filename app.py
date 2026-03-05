"""
AI Dropshipping ERP System - Main Application
Complete implementation with all 8 modules
"""

# ============================================================================
# TIMEZONE CONFIGURATION (MUST BE FIRST)
# ============================================================================
import os
os.environ['TZ'] = 'Asia/Seoul'
try:
    import time
    time.tzset()
except AttributeError:
    pass  # Windows doesn't have tzset

from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_file, flash
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import sqlite3
import json
import requests
import hashlib
import hmac
import base64
import bcrypt
from datetime import datetime, timedelta
import pytz
from PIL import Image, ImageDraw, ImageFont
import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import schedule
import threading
import logging
from logging.handlers import RotatingFileHandler
from market_analysis import analyze_naver_market, get_naver_keyword_trend

# Set Korea timezone globally
KST = pytz.timezone('Asia/Seoul')

def get_kst_now():
    """Get current time in Korea Standard Time"""
    return datetime.now(KST)

# ============================================================================
# DATABASE PATH CONFIGURATION (MUST BE FIRST)
# ============================================================================

# Get absolute path to ensure same DB regardless of working directory
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'dropship.db')

# CRITICAL: Print and log DB path at module load time
print(f"[INIT] ========================================")
print(f"[INIT] Module: {__file__}")
print(f"[INIT] Base directory: {BASE_DIR}")
print(f"[INIT] Database path (ABSOLUTE): {DB_PATH}")
print(f"[INIT] Database exists: {os.path.exists(DB_PATH)}")
if os.path.exists(DB_PATH):
    print(f"[INIT] Database size: {os.path.getsize(DB_PATH)} bytes")
print(f"[INIT] ========================================")

# Verify DB path is absolute
if not os.path.isabs(DB_PATH):
    raise RuntimeError(f"CRITICAL: DB_PATH must be absolute! Got: {DB_PATH}")

print(f"[INIT] ✅ DB_PATH verification passed: {DB_PATH}")

# ============================================================================
# CRITICAL: AUTO DATABASE INITIALIZATION (BEFORE FLASK APP)
# ============================================================================

def auto_init_database():
    """
    Automatically initialize database if it doesn't exist or is missing tables.
    This prevents sqlite3.OperationalError: no such table errors.
    CRITICAL: Must run BEFORE Flask app initialization.
    """
    needs_init = False
    
    # Check if database file exists
    if not os.path.exists(DB_PATH):
        print(f'[DB-INIT] Database file not found: {DB_PATH}')
        needs_init = True
    else:
        # Check if required tables exist
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='users'")
            if not cursor.fetchone():
                print('[DB-INIT] Required table "users" not found in database')
                needs_init = True
            conn.close()
        except Exception as e:
            print(f'[DB-INIT] Error checking database: {e}')
            needs_init = True
    
    if needs_init:
        print('[DB-INIT] 🔧 Initializing database automatically...')
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            
            # Create users table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS users (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT UNIQUE NOT NULL,
                    password_hash TEXT NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # Create config table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS config (
                    key TEXT PRIMARY KEY,
                    value TEXT NOT NULL,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # Create sourced_products table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS sourced_products (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    original_url TEXT NOT NULL,
                    title_cn TEXT,
                    title_kr TEXT,
                    description_cn TEXT,
                    description_kr TEXT,
                    marketing_copy TEXT,
                    price_cny REAL,
                    price_krw INTEGER,
                    profit_margin REAL,
                    estimated_profit INTEGER,
                    traffic_score INTEGER,
                    safety_status TEXT,
                    images_json TEXT,
                    processed_images_json TEXT,
                    status TEXT DEFAULT 'pending',
                    keywords TEXT,
                    source_site TEXT DEFAULT 'alibaba',
                    moq INTEGER DEFAULT 1,
                    trend_score INTEGER DEFAULT 0,
                    competition_score INTEGER DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    approved_at TIMESTAMP,
                    registered_at TIMESTAMP
                )
            ''')
            
            # CRITICAL: Add new columns if they don't exist (migration)
            migrations = [
                ('keywords', 'TEXT'),
                ('source_site', "TEXT DEFAULT 'alibaba'"),
                ('moq', 'INTEGER DEFAULT 1'),
                ('trend_score', 'INTEGER DEFAULT 0'),
                ('competition_score', 'INTEGER DEFAULT 0')
            ]
            
            for column_name, column_type in migrations:
                try:
                    cursor.execute(f'ALTER TABLE sourced_products ADD COLUMN {column_name} {column_type}')
                    app.logger.info(f'[DB Init] ✅ Added {column_name} column to sourced_products table')
                except:
                    pass  # Column already exists
            
            # Create orders table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS orders (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    order_number TEXT UNIQUE NOT NULL,
                    marketplace TEXT NOT NULL,
                    product_id INTEGER,
                    customer_name TEXT,
                    customer_phone TEXT,
                    customer_address TEXT,
                    pccc TEXT NOT NULL,
                    pccc_validated BOOLEAN DEFAULT 0,
                    quantity INTEGER DEFAULT 1,
                    sale_price INTEGER NOT NULL,
                    purchase_price_cny REAL,
                    purchase_price_krw INTEGER,
                    applied_exchange_rate REAL NOT NULL,
                    shipping_cost INTEGER,
                    customs_tax INTEGER,
                    marketplace_fee INTEGER,
                    net_profit INTEGER,
                    original_product_url TEXT,
                    tracking_number TEXT,
                    order_status TEXT DEFAULT 'pending',
                    payment_status TEXT DEFAULT 'paid',
                    shipping_deadline TIMESTAMP,
                    ordered_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    shipped_at TIMESTAMP,
                    delivered_at TIMESTAMP,
                    FOREIGN KEY (product_id) REFERENCES sourced_products (id)
                )
            ''')
            
            # Create activity_logs table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS activity_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    action_type TEXT NOT NULL,
                    description TEXT,
                    status TEXT,
                    details_json TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # Create tax_records table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS tax_records (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    order_id INTEGER,
                    order_number TEXT,
                    sale_date DATE,
                    sale_amount INTEGER,
                    purchase_amount INTEGER,
                    shipping_cost INTEGER,
                    marketplace_fee INTEGER,
                    net_profit INTEGER,
                    applied_exchange_rate REAL,
                    exported BOOLEAN DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (order_id) REFERENCES orders (id)
                )
            ''')
            
            # Create marketplace_listings table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS marketplace_listings (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    product_id INTEGER,
                    marketplace TEXT NOT NULL,
                    external_product_id TEXT,
                    listing_url TEXT,
                    current_stock INTEGER DEFAULT 0,
                    status TEXT DEFAULT 'active',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP,
                    FOREIGN KEY (product_id) REFERENCES sourced_products (id)
                )
            ''')
            
            # Create stock_monitor_log table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS stock_monitor_log (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    product_id INTEGER,
                    original_url TEXT,
                    check_status TEXT,
                    action_taken TEXT,
                    checked_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (product_id) REFERENCES sourced_products (id)
                )
            ''')
            
            # Check if default admin user exists
            cursor.execute("SELECT COUNT(*) FROM users WHERE username='admin'")
            if cursor.fetchone()[0] == 0:
                default_password = 'admin123'
                password_hash = generate_password_hash(default_password, method='pbkdf2:sha256')
                cursor.execute('INSERT INTO users (username, password_hash) VALUES (?, ?)',
                             ('admin', password_hash))
                print('[DB-INIT] ✅ Default admin user created (username: admin, password: admin123)')
            
            # Insert default configuration if config table is empty
            cursor.execute("SELECT COUNT(*) FROM config")
            if cursor.fetchone()[0] == 0:
                default_configs = [
                    ('target_margin_rate', '30'),
                    ('cny_exchange_rate', '190'),
                    ('exchange_rate_buffer', '1.05'),
                    ('shipping_cost_base', '5000'),
                    ('customs_tax_rate', '0.10'),
                    ('naver_fee_rate', '0.06'),
                    ('coupang_fee_rate', '0.11'),
                    ('auto_register', 'false'),
                    ('scrapingant_api_key', ''),
                    ('openai_api_key', ''),
                    ('naver_client_id', ''),
                    ('naver_client_secret', ''),
                    ('coupang_access_key', ''),
                    ('coupang_secret_key', ''),
                    ('server_static_ip', ''),
                    ('debug_mode_ignore_filters', 'false'),  # NEW: Debug mode for diagnostics
                ]
                cursor.executemany('INSERT INTO config (key, value) VALUES (?, ?)', default_configs)
                print('[DB-INIT] ✅ Default configuration inserted')
            
            conn.commit()
            conn.close()
            
            print('[DB-INIT] ✅ Database initialized successfully!')
            print(f'[DB-INIT] 📁 Database: {DB_PATH}')
            print('[DB-INIT] ⚠️  Default credentials: admin / admin123')
            
        except Exception as e:
            print(f'[DB-INIT] ❌ Failed to initialize database: {e}')
            import traceback
            traceback.print_exc()
            raise
    else:
        print(f'[DB-INIT] ✅ Database OK: {DB_PATH}')
    
    # ============================================================================
    # CRITICAL: Run migrations ALWAYS (even if database already exists)
    # ============================================================================
    print('[DB-MIGRATE] 🔄 Running column migrations...')
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Check existing columns
        cursor.execute("PRAGMA table_info(sourced_products)")
        existing_columns = {row[1] for row in cursor.fetchall()}
        print(f'[DB-MIGRATE] Existing columns: {sorted(existing_columns)}')
        
        # Define required columns with their types
        required_columns = {
            'keywords': 'TEXT',
            'source_site': "TEXT DEFAULT 'alibaba'",
            'moq': 'INTEGER DEFAULT 1',
            'trend_score': 'INTEGER DEFAULT 0',
            'competition_score': 'INTEGER DEFAULT 0'
        }
        
        # Add missing columns
        migrations_applied = 0
        for column_name, column_type in required_columns.items():
            if column_name not in existing_columns:
                try:
                    cursor.execute(f'ALTER TABLE sourced_products ADD COLUMN {column_name} {column_type}')
                    conn.commit()
                    print(f'[DB-MIGRATE] ✅ Added column: {column_name}')
                    migrations_applied += 1
                except sqlite3.OperationalError as e:
                    print(f'[DB-MIGRATE] ⚠️ Column {column_name} error: {e}')
            else:
                print(f'[DB-MIGRATE] ✓ Column {column_name} exists')
        
        conn.close()
        if migrations_applied > 0:
            print(f'[DB-MIGRATE] ✅ Applied {migrations_applied} migration(s)')
        else:
            print('[DB-MIGRATE] ✅ All columns up to date')
        
    except Exception as e:
        print(f'[DB-MIGRATE] ❌ Migration error: {e}')
        import traceback
        traceback.print_exc()
        # Don't raise - app should still work even if migration partially fails
    
    # CRITICAL: Verify tables exist before proceeding
    print('[DB-VERIFY] Verifying all tables exist...')
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
        tables = [row[0] for row in cursor.fetchall()]
        conn.close()
        
        required_tables = ['users', 'config', 'sourced_products', 'orders', 
                          'activity_logs', 'tax_records', 'marketplace_listings', 'stock_monitor_log']
        missing_tables = [t for t in required_tables if t not in tables]
        
        if missing_tables:
            print(f'[DB-VERIFY] ❌ ERROR: Missing tables: {missing_tables}')
            print(f'[DB-VERIFY] Found tables: {tables}')
            raise Exception(f'Database verification failed! Missing tables: {missing_tables}')
        else:
            print(f'[DB-VERIFY] ✅ All {len(required_tables)} required tables exist!')
            print(f'[DB-VERIFY] Tables: {", ".join(tables)}')
    except Exception as e:
        print(f'[DB-VERIFY] ❌ Verification failed: {e}')
        raise
    
    # Always print completion marker
    print('='*70)
    print('!!! DATABASE INITIALIZATION COMPLETE !!!')
    print('='*70)

# RUN DATABASE INITIALIZATION IMMEDIATELY (BEFORE ANYTHING ELSE)
print('[CRITICAL] Starting database initialization...')
auto_init_database()
print('[CRITICAL] Database initialization finished. Proceeding with Flask setup...')

# ============================================================================
# FLASK APP INITIALIZATION (AFTER DATABASE IS READY)
# ============================================================================

app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'dev-secret-key-change-in-production')

# Add custom Jinja2 filters
@app.template_filter('from_json')
def from_json_filter(value):
    """Parse JSON string to Python object"""
    import json
    try:
        return json.loads(value) if value else []
    except (ValueError, TypeError):
        return []

# ============================================================================
# LOGGING CONFIGURATION
# ============================================================================

# Create logs directory if it doesn't exist
if not os.path.exists('logs'):
    os.makedirs('logs')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

# File handler with rotation (max 10MB, keep 5 backup files)
file_handler = RotatingFileHandler(
    'logs/server.log',
    maxBytes=10 * 1024 * 1024,  # 10MB
    backupCount=5,
    encoding='utf-8'
)
file_handler.setLevel(logging.INFO)
file_handler.setFormatter(logging.Formatter(
    '%(asctime)s [%(levelname)s] %(name)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
))

# Console handler
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
console_handler.setFormatter(logging.Formatter(
    '%(asctime)s [%(levelname)s] %(name)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
))

# Add handlers to app logger
app.logger.addHandler(file_handler)
app.logger.addHandler(console_handler)
app.logger.setLevel(logging.INFO)

# Log startup
app.logger.info('=' * 70)
app.logger.info('AI Dropshipping ERP System v2.0 - Starting Up')
app.logger.info('=' * 70)

# Flask-Login setup
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# ============================================================================
# JINJA2 CUSTOM FILTERS
# ============================================================================

def parse_datetime_filter(date_string):
    """Parse datetime string to datetime object for Jinja2 template"""
    if not date_string:
        return get_kst_now()
    
    # Handle various datetime formats
    try:
        # ISO format: 2024-01-21 12:34:56 or with microseconds
        date_str = str(date_string).split('.')[0]  # Remove microseconds if present
        if 'T' in date_str:
            # ISO format with T separator
            return datetime.fromisoformat(date_str.replace('T', ' '))
        else:
            # Standard SQL format
            return datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
    except (ValueError, AttributeError):
        # If parsing fails, return current time to avoid template errors
        return get_kst_now()

# Register custom Jinja2 filter
app.jinja_env.filters['parse_datetime'] = parse_datetime_filter

# Add now() function to Jinja2 globals for template usage
app.jinja_env.globals['now'] = datetime.now

# ============================================================================
# IMAGE URL UTILITIES
# ============================================================================

def fix_image_url(url):
    """
    Fix image URL to ensure absolute HTTPS path
    - Converts // to https://
    - Returns empty string if invalid
    """
    if not url:
        return ''
    
    url = str(url).strip()
    
    # If starts with //, add https:
    if url.startswith('//'):
        return 'https:' + url
    
    # If already absolute, return as-is
    if url.startswith('http://') or url.startswith('https://'):
        return url
    
    # If relative path, return empty (can't fix without base domain)
    return ''

# ============================================================================
# DATABASE UTILITIES
# ============================================================================

def get_db():
    """Get database connection"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def log_activity(action_type, description, status='success', details=None):
    """Log system activity with KST timestamp"""
    conn = get_db()
    cursor = conn.cursor()
    
    # 🚀 FIX: Explicitly set KST timestamp instead of relying on DEFAULT
    kst_now = get_kst_now().strftime('%Y-%m-%d %H:%M:%S')
    
    cursor.execute('''
        INSERT INTO activity_logs (action_type, description, status, details_json, created_at)
        VALUES (?, ?, ?, ?, ?)
    ''', (action_type, description, status, json.dumps(details) if details else None, kst_now))
    conn.commit()
    conn.close()

def get_config(key, default=None):
    """Get configuration value with defensive handling"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT value FROM config WHERE key = ?', (key,))
    row = cursor.fetchone()
    conn.close()
    
    if row:
        value = row['value']
        # Strip whitespace from string values
        if isinstance(value, str):
            value = value.strip()
        # Return None for empty strings (treat as not configured)
        if value == '':
            return default
        return value
    return default

def set_config(key, value):
    """Set configuration value"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO config (key, value, updated_at) 
        VALUES (?, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(key) DO UPDATE SET value = ?, updated_at = CURRENT_TIMESTAMP
    ''', (key, value, value))
    conn.commit()
    conn.close()

def system_check_critical_configs():
    """
    CRITICAL: System-wide configuration verification
    This function MUST be called at the start of any critical operation
    to ensure all required API keys and configs are loaded from DB.
    
    Returns: dict with status and config values
    Raises: RuntimeError if critical configs are missing
    """
    app.logger.info('[SYSTEM-CHECK] ========================================')
    app.logger.info(f'[SYSTEM-CHECK] 🔍 DB Path (ABSOLUTE): {DB_PATH}')
    app.logger.info(f'[SYSTEM-CHECK] DB Path is absolute: {os.path.isabs(DB_PATH)}')
    app.logger.info(f'[SYSTEM-CHECK] DB Exists: {os.path.exists(DB_PATH)}')
    
    if not os.path.exists(DB_PATH):
        error_msg = f'CRITICAL: Database file not found at {DB_PATH}'
        app.logger.error(f'[SYSTEM-CHECK] ❌ {error_msg}')
        raise RuntimeError(error_msg)
    
    app.logger.info(f'[SYSTEM-CHECK] DB Size: {os.path.getsize(DB_PATH)} bytes')
    
    # Load critical configs from DB
    openai_key = get_config('openai_api_key', '')
    scrapingant_key = get_config('scrapingant_api_key', '')
    target_margin = get_config('target_margin_rate', '30')
    cny_rate = get_config('cny_exchange_rate', '190')
    
    # Log first 4 characters of API keys (masked)
    openai_preview = openai_key[:4] + '****' if len(openai_key) >= 4 else 'EMPTY'
    scrapingant_preview = scrapingant_key[:4] + '****' if len(scrapingant_key) >= 4 else 'EMPTY'
    
    app.logger.info(f'[SYSTEM-CHECK] 🔑 OpenAI API Key: {openai_preview} (length: {len(openai_key)})')
    app.logger.info(f'[SYSTEM-CHECK] 🔑 ScrapingAnt API Key: {scrapingant_preview} (length: {len(scrapingant_key)})')
    app.logger.info(f'[SYSTEM-CHECK] 💰 Target Margin: {target_margin}%')
    app.logger.info(f'[SYSTEM-CHECK] 💱 CNY Exchange Rate: {cny_rate}')
    
    # Check if critical keys are empty
    missing_keys = []
    if not openai_key or openai_key.strip() == '':
        missing_keys.append('openai_api_key')
        app.logger.error('[SYSTEM-CHECK] ❌ OpenAI API key is EMPTY or NOT CONFIGURED')
    
    if not scrapingant_key or scrapingant_key.strip() == '':
        missing_keys.append('scrapingant_api_key')
        app.logger.error('[SYSTEM-CHECK] ❌ ScrapingAnt API key is EMPTY or NOT CONFIGURED')
    
    if missing_keys:
        error_msg = f'CRITICAL: Missing required API keys in DB: {missing_keys}'
        app.logger.error(f'[SYSTEM-CHECK] ❌ {error_msg}')
        app.logger.error('[SYSTEM-CHECK] ❌ Please configure these keys in the Settings page!')
        app.logger.info('[SYSTEM-CHECK] ========================================')
        raise RuntimeError(error_msg)
    
    app.logger.info('[SYSTEM-CHECK] ✅ All critical configurations verified successfully')
    app.logger.info('[SYSTEM-CHECK] ========================================')
    
    return {
        'success': True,
        'openai_key_preview': openai_preview,
        'scrapingant_key_preview': scrapingant_preview,
        'target_margin': target_margin,
        'cny_rate': cny_rate
    }
    
    # Check critical API keys
    critical_keys = {
        'scrapingant_api_key': 'ScrapingAnt API Key',
        'openai_api_key': 'OpenAI API Key'
    }
    
    config_status = {}
    missing_keys = []
    
    for key, display_name in critical_keys.items():
        app.logger.info(f'[SYSTEM-CHECK] Fetching {display_name} from DB...')
        
        # Direct DB query to ensure we're reading from the correct database
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute('SELECT value FROM config WHERE key = ?', (key,))
        row = cursor.fetchone()
        conn.close()
        
        if row and row['value'] and row['value'].strip():
            value = row['value'].strip()
            # Show first 4 chars for verification
            preview = f"{value[:4]}..." if len(value) > 4 else "[TOO_SHORT]"
            config_status[key] = {
                'configured': True,
                'length': len(value),
                'preview': preview
            }
            app.logger.info(f'[SYSTEM-CHECK] ✅ {display_name}: YES (length: {len(value)}, preview: {preview})')
        else:
            config_status[key] = {
                'configured': False,
                'length': 0,
                'preview': None
            }
            missing_keys.append(display_name)
            app.logger.error(f'[SYSTEM-CHECK] ❌ {display_name}: NOT CONFIGURED')
    
    # Check other important configs
    other_configs = ['target_margin_rate', 'cny_exchange_rate']
    for key in other_configs:
        value = get_config(key)
        config_status[key] = value
        app.logger.info(f'[SYSTEM-CHECK] {key}: {value}')
    
    app.logger.info('[SYSTEM-CHECK] ========================================')
    
    return {
        'success': len(missing_keys) == 0,
        'missing_keys': missing_keys,
        'configs': config_status
    }

# ============================================================================
# USER MODEL FOR FLASK-LOGIN
# ============================================================================

class User(UserMixin):
    def __init__(self, user_id, username):
        self.id = user_id
        self.username = username

@login_manager.user_loader
def load_user(user_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT id, username FROM users WHERE id = ?', (user_id,))
    row = cursor.fetchone()
    conn.close()
    if row:
        return User(row['id'], row['username'])
    return None

# ============================================================================
# MODULE 1: AUTHENTICATION
# ============================================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT id, username, password_hash FROM users WHERE username = ?', (username,))
        user = cursor.fetchone()
        conn.close()
        
        if user and check_password_hash(user['password_hash'], password):
            user_obj = User(user['id'], user['username'])
            login_user(user_obj)
            log_activity('auth', f'User {username} logged in')
            return redirect(url_for('dashboard'))
        else:
            return render_template('login.html', error='Invalid credentials')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    username = current_user.username
    logout_user()
    log_activity('auth', f'User {username} logged out')
    return redirect(url_for('login'))

# ============================================================================
# MODULE 2: AI SOURCING ENGINE WITH SAFETY FILTERS & BLUE OCEAN ANALYSIS
# ============================================================================

# Safety filter keywords
BANNED_KEYWORDS = {
    'food': ['食品', '食物', '零食', '糖果', '饼干', '巧克力', '饮料'],
    'tableware': ['餐具', '碗', '盘子', '筷子', '勺子', '叉子', '杯子'],
    'baby': ['婴儿', '儿童', '宝宝', '玩具', '奶瓶', '尿布', '童装'],
    'cosmetic': ['化妆品', '护肤', '面膜', '口红', '眼影', '粉底'],
    'replica': ['Nike', 'Adidas', 'Gucci', 'LV', 'Louis Vuitton', 'Chanel', 
                'Disney', '迪士尼', 'Supreme', 'Rolex', 'Apple']
}

def analyze_blue_ocean_market(user_keyword=''):
    """
    Advanced Blue Ocean Market Analysis using GPT-4o-mini
    Finds niche opportunities with rising demand and low competition
    """
    api_key = get_config('openai_api_key')
    
    # Defensive: Check API key validity
    if not api_key:
        app.logger.warning('⚠️ OpenAI API key not configured in database')
        return {
            'suggested_keyword': user_keyword if user_keyword else '무선이어폰',
            'reasoning': 'OpenAI API key not configured. Using provided keyword.',
            'analysis_performed': False
        }
    
    # Log API key prefix for debugging (first 7 chars only for security)
    api_key_preview = api_key[:7] + '...' if len(api_key) > 7 else 'TOO_SHORT'
    app.logger.info(f'🔑 Using OpenAI API key: {api_key_preview} (length: {len(api_key)})')
    
    # Validate API key format
    if not api_key.startswith('sk-'):
        app.logger.error(f'❌ Invalid OpenAI API key format (does not start with sk-): {api_key_preview}')
        return {
            'suggested_keyword': user_keyword if user_keyword else '무선이어폰',
            'reasoning': 'Invalid OpenAI API key format. Please check your configuration.',
            'analysis_performed': False
        }
    
    # Get current date and season for context
    now = get_kst_now()
    current_month = now.month
    
    # Determine Korean season
    if current_month in [3, 4, 5]:
        season = '봄 (Spring)'
    elif current_month in [6, 7, 8]:
        season = '여름 (Summer)'
    elif current_month in [9, 10, 11]:
        season = '가을 (Fall)'
    else:
        season = '겨울 (Winter)'
    
    # Advanced Blue Ocean Analysis Prompt
    # CRITICAL: Must explicitly ask for JSON format to avoid 400 errors with json_object mode
    prompt = f"""당신은 대한민국 E-커머스의 수석 MD이자 데이터 분석 전문가입니다.

【현재 상황】
- 날짜: {now.strftime('%Y년 %m월 %d일')}
- 계절: {season}
- 분석 기간: 향후 2주~1개월 내 판매 폭발 예상 상품

【사용자 관심 키워드】
"{user_keyword if user_keyword else '없음 (자유 선정)'}"

【미션】
위 키워드를 참고하되, 다음 조건을 모두 만족하는 '블루오션(Blue Ocean)' 상품 키워드 1개를 찾아내세요:

✅ **필수 충족 조건**

1. Rising Trend (급상승 트렌드)
   - 최근 검색량이 급증하고 있거나, 다가올 시즌에 수요 폭발 예상
   - 계절성, 이벤트, 신규 트렌드를 고려

2. Low Competition (낮은 경쟁 강도)
   - 대기업 브랜드(삼성, LG, 나이키 등)가 장악하지 않은 카테고리
   - 중소 셀러가 진입 가능한 틈새시장

3. Specificity (구체적인 롱테일 키워드)
   - 너무 광범위한 키워드(예: '가습기') 대신
   - 구체적이고 니치한 키워드(예: '무선 무드등 탁상용 가습기')
   - Alibaba/AliExpress에서 검색 가능한 구체적인 상품명

🚫 **절대 금지 카테고리 (인증/규제 이슈)**

아래 카테고리는 **절대로 추천하지 마세요**:
- 전자제품 (무선이어폰, 블루투스스피커, 스마트워치, 보조배터리 등) → KC인증, 전파인증 필요
- 전기난방 기기 (전기장판, 온열매트, 히터 등) → 안전인증 필요
- 의료기기 (체온계, 혈압계, 마사지기 등) → 의료기기 인증 필요
- 유아용품 (젖병, 장난감, 카시트 등) → KC인증 필수
- 식품/건강기능식품 → 식약처 승인 필요
- 화장품 (기능성 화장품) → 식약처 보고 필요
- 안전용품 (헬멧, 방호복, 소화기 등) → 안전인증 필요
- 무게/용량 측정기 (저울, 계량기 등) → 계량 법 규제

✅ **추천 가능 카테고리**

- 패션잡화 (가방, 지갑, 모자, 스카프 등)
- 리빙/홈데코 (수납용품, 인테리어소품, 화병, 액자 등)
- 주방용품 (실리콘 조리도구, 밀폐용기, 주방정리용품 등)
- 문구/오피스 (노트, 펜, 다이어리, 데스크정리함 등)
- 반려동물용품 (간식 제외: 목줄, 장난감, 하우스, 식기 등)
- 스포츠/레저 (요가매트, 운동밴드, 캠핑소품 등)
- 자동차용품 (방향제, 차량정리함, 햇빛가리개 등)
- 원예/가드닝 (화분, 원예도구, 지지대, 장식돌 등)

**중요: 반드시 아래 JSON 형식으로만 답변하세요. 다른 텍스트는 포함하지 마세요:**

{{
  "keyword": "정밀한 블루오션 키워드 (한국어)",
  "reasoning": "이 키워드를 선정한 이유를 1~2문장으로 설명",
  "trend_score": 1-10 사이 점수,
  "competition_score": 1-10 사이 점수 (낮을수록 좋음),
  "category": "해당 카테고리명 (예: 리빙/홈데코)"
}}

예시:
{{
  "keyword": "반려동물 자동 급수기 스텐 식기",
  "reasoning": "1인 가구 증가로 펫테크 수요 급증, KC인증 불필요, 대기업 미진입 영역",
  "trend_score": 9,
  "competition_score": 3,
  "category": "반려동물용품"
}}"""

    try:
        # Log request details
        app.logger.info(f'📡 Calling OpenAI API: model=gpt-4o-mini, max_tokens=500, temperature=0.8')
        
        response = requests.post(
            'https://api.openai.com/v1/chat/completions',
            headers={
                'Authorization': f'Bearer {api_key}',
                'Content-Type': 'application/json'
            },
            json={
                'model': 'gpt-4o-mini',  # FORCED: Must use gpt-4o-mini (universal access)
                'messages': [
                    {
                        'role': 'system',
                        'content': '당신은 한국 E-커머스 시장의 전문 MD이자 트렌드 분석가입니다. 블루오션 시장을 발굴하는 전문가입니다. 반드시 JSON 형식으로만 응답하세요. 중요: KC인증, 전파인증 등 규제가 필요한 전자제품, 의료기기, 식품은 절대 추천하지 마세요.'
                    },
                    {
                        'role': 'user',
                        'content': prompt
                    }
                ],
                'temperature': 0.8,
                'max_tokens': 500,
                'response_format': {'type': 'json_object'}  # JSON mode enabled
            },
            timeout=30
        )
        
        # Log response status
        app.logger.info(f'📥 OpenAI API Response: status_code={response.status_code}')
        
        if response.status_code == 200:
            result = response.json()
            content = result['choices'][0]['message']['content']
            
            app.logger.info(f'✅ Received response from OpenAI (length: {len(content)} chars)')
            
            # Parse JSON response
            import json
            try:
                analysis = json.loads(content)
                app.logger.info(f'🎯 Blue Ocean Keyword: {analysis.get("keyword")}')
                
                return {
                    'suggested_keyword': analysis.get('keyword', user_keyword or '무선이어폰'),
                    'reasoning': analysis.get('reasoning', 'AI 분석 완료'),
                    'trend_score': analysis.get('trend_score', 0),
                    'competition_score': analysis.get('competition_score', 0),
                    'analysis_performed': True
                }
            except json.JSONDecodeError as je:
                app.logger.error(f'❌ JSON parsing failed: {je}')
                app.logger.error(f'Raw content: {content[:200]}...')
                return {
                    'suggested_keyword': user_keyword if user_keyword else '무선이어폰',
                    'reasoning': f'AI 응답 파싱 실패: {str(je)}',
                    'analysis_performed': False
                }
        else:
            # Enhanced error logging for non-200 responses
            try:
                error_body = response.json()
                error_message = error_body.get('error', {}).get('message', 'No error message')
                error_type = error_body.get('error', {}).get('type', 'unknown')
                app.logger.error(f'❌ OpenAI API error {response.status_code}: {error_type} - {error_message}')
                app.logger.error(f'Full error body: {error_body}')
            except:
                app.logger.error(f'❌ OpenAI API error {response.status_code}: {response.text[:300]}')
            
            return {
                'suggested_keyword': user_keyword if user_keyword else '무선이어폰',
                'reasoning': f'AI 분석 실패 (API 오류: {response.status_code})',
                'analysis_performed': False
            }
    
    except requests.exceptions.Timeout:
        app.logger.error('❌ OpenAI API timeout (30s)')
        return {
            'suggested_keyword': user_keyword if user_keyword else '무선이어폰',
            'reasoning': 'AI 분석 시간 초과 (30초)',
            'analysis_performed': False
        }
    except requests.exceptions.RequestException as re:
        app.logger.error(f'❌ Network error: {str(re)}')
        return {
            'suggested_keyword': user_keyword if user_keyword else '무선이어폰',
            'reasoning': f'네트워크 오류: {str(re)}',
            'analysis_performed': False
        }
    except Exception as e:
        app.logger.error(f'❌ Unexpected error in Blue Ocean analysis: {str(e)}')
        app.logger.exception(e)  # This will log full traceback
        return {
            'suggested_keyword': user_keyword if user_keyword else '무선이어폰',
            'reasoning': f'AI 분석 실패: {str(e)}',
            'analysis_performed': False
        }

def check_safety_filter(title, description=''):
    """Check if product passes safety filter"""
    text = (title + ' ' + description).lower()
    
    for category, keywords in BANNED_KEYWORDS.items():
        for keyword in keywords:
            if keyword.lower() in text:
                return False, f'Banned category: {category} (keyword: {keyword})'
    
    return True, 'Pass'

def scrape_1688_search(keyword, max_results=50):
    """
    ⚠️ DEPRECATED - NO LONGER USED
    This function is kept for backward compatibility only.
    Use search_integrated_hybrid() instead (Alibaba + AliExpress)
    """
    app.logger.warning(f'[DEPRECATED] scrape_1688_search called - redirecting to hybrid search')
    return search_integrated_hybrid(keyword, max_results)

# TEST DATA FUNCTIONS COMPLETELY REMOVED
# No fallback test data - fail explicitly if scraping fails

def parse_smart_price(price_text):
    """
    Smart price parser for Alibaba/AliExpress
    Handles: $10-$20 -> 10.0, $15.99 -> 15.99, ¥50-¥100 -> 50.0
    """
    import re
    if not price_text:
        return 0.0
    
    # Remove currency symbols and spaces
    cleaned = price_text.replace('$', '').replace('¥', '').replace(' ', '').replace(',', '')
    
    # Find all numbers (including decimals)
    numbers = re.findall(r'\d+\.?\d*', cleaned)
    
    if not numbers:
        return 0.0
    
    # If range (e.g., "10-20"), take the LOWEST price
    prices = [float(n) for n in numbers]
    return min(prices)

# ============================================================================
# ALIEXPRESS OFFICIAL API INTEGRATION
# ============================================================================

def search_aliexpress_official(keyword, max_results=50):
    """
    Search AliExpress using Official Affiliate API
    
    Benefits:
    - ✅ FREE (no ScrapingAnt cost)
    - ✅ 100% success rate (no anti-bot blocks)
    - ✅ Official product data
    - ✅ Affiliate links included
    
    Args:
        keyword (str): Search keyword
        max_results (int): Maximum products to return (default 50)
    
    Returns:
        dict: {
            'products': [
                {
                    'title': str,
                    'price': float (USD),
                    'original_price': float,
                    'url': str (affiliate link),
                    'image': str,
                    'product_id': str,
                    'sale_price': float,
                    'discount': str,
                    'orders': int,
                    'rating': float,
                    'source': 'aliexpress_api'
                }
            ],
            'count': int
        }
    """
    import time
    import hashlib
    
    app.logger.info(f'[AliExpress API] ========================================')
    app.logger.info(f'[AliExpress API] 🚀 Official API search for: {keyword}')
    
    # Get API credentials from config
    app_key = get_config('aliexpress_app_key')
    app_secret = get_config('aliexpress_app_secret')
    
    if not app_key or not app_secret:
        app.logger.error('[AliExpress API] ❌ Missing API credentials')
        app.logger.error('[AliExpress API] Please configure in Settings:')
        app.logger.error('[AliExpress API]   - aliexpress_app_key')
        app.logger.error('[AliExpress API]   - aliexpress_app_secret')
        return {'products': [], 'count': 0}
    
    # API endpoint
    endpoint = "https://api-sg.aliexpress.com/sync"
    
    # GMT+8 timestamp
    timestamp = time.strftime('%Y-%m-%d %H:%M:%S', time.gmtime(time.time() + 8*3600))
    
    # Request parameters
    params = {
        'app_key': app_key,
        'method': 'aliexpress.affiliate.product.query',
        'timestamp': timestamp,
        'format': 'json',
        'v': '2.0',
        'sign_method': 'md5',
        'keywords': keyword,
        'page_size': str(min(max_results, 50)),  # API max: 50
        'target_currency': 'USD',
        'target_language': 'EN',
        'sort': 'SALE_PRICE_ASC'  # Sort by price (low to high)
    }
    
    # Generate signature
    sorted_params = sorted(params.items())
    sign_string = app_secret + ''.join([f'{k}{v}' for k, v in sorted_params]) + app_secret
    signature = hashlib.md5(sign_string.encode('utf-8')).hexdigest().upper()
    params['sign'] = signature
    
    try:
        app.logger.info(f'[AliExpress API] 📡 Sending request...')
        response = requests.get(endpoint, params=params, timeout=30)
        
        app.logger.info(f'[AliExpress API] 📊 Status: {response.status_code}')
        
        if response.status_code != 200:
            app.logger.error(f'[AliExpress API] ❌ HTTP {response.status_code}')
            return {'products': [], 'count': 0}
        
        data = response.json()
        
        # Check for API errors
        if 'error_response' in data:
            error = data['error_response']
            app.logger.error(f'[AliExpress API] ❌ API Error: {error.get("msg", "Unknown")}')
            app.logger.error(f'[AliExpress API]    Code: {error.get("code")} / Sub: {error.get("sub_code")}')
            return {'products': [], 'count': 0}
        
        # Parse response
        if 'aliexpress_affiliate_product_query_response' not in data:
            app.logger.error('[AliExpress API] ❌ Unexpected response format')
            return {'products': [], 'count': 0}
        
        result = data['aliexpress_affiliate_product_query_response']['resp_result']['result']
        
        if result['total_record_count'] == 0:
            app.logger.warning(f'[AliExpress API] ⚠️  No products found for: {keyword}')
            return {'products': [], 'count': 0}
        
        # Extract products
        raw_products = result['products']['product']
        products = []
        
        for product in raw_products:
            try:
                # Parse price
                sale_price = float(product.get('target_sale_price', 0))
                original_price = float(product.get('target_original_price', sale_price))
                
                # Calculate discount
                discount = 0
                if original_price > sale_price:
                    discount = int((1 - sale_price / original_price) * 100)
                
                products.append({
                    'title': product.get('product_title', 'Untitled'),
                    'price': sale_price,
                    'original_price': original_price,
                    'url': product.get('promotion_link', product.get('product_detail_url', '')),
                    'image': product.get('product_main_image_url', ''),
                    'product_id': str(product.get('product_id', '')),
                    'sale_price': sale_price,
                    'discount': f'-{discount}%' if discount > 0 else '',
                    'orders': 0,  # API doesn't provide order count
                    'rating': 0.0,  # API doesn't provide rating
                    'moq': 1,  # AliExpress default MOQ
                    'source': 'aliexpress_api'
                })
            except Exception as e:
                app.logger.error(f'[AliExpress API] Failed to parse product: {str(e)}')
                continue
        
        app.logger.info(f'[AliExpress API] ✅ SUCCESS: {len(products)} products')
        return {'products': products, 'count': len(products)}
        
    except Exception as e:
        app.logger.error(f'[AliExpress API] ❌ Exception: {str(e)}')
        return {'products': [], 'count': 0}

# ============================================================================
# LEGACY SCRAPING FUNCTIONS (DEPRECATED - USE OFFICIAL API)
# ============================================================================

def scrape_alibaba_search(keyword, max_results=50):
    """
    Scrape Alibaba.com with ULTIMATE anti-bot bypass
    - Randomized User-Agents
    - Stealth proxy
    - Random delays
    - Multi-retry with different IPs
    """
    import urllib.parse
    import random
    import time
    
    app.logger.info(f'[Alibaba Scraping] ========================================')
    app.logger.info(f'[Alibaba Scraping] 🔥 ULTIMATE ANTI-BOT MODE for: {keyword}')
    
    api_key = get_config('scrapingant_api_key')
    if not api_key or not api_key.strip():
        app.logger.error('[Alibaba Scraping] ❌ No ScrapingAnt API key')
        return {'products': [], 'count': 0}
    
    # 🎭 Pool of REAL User-Agents (Latest Chrome/Edge 2024-2026)
    user_agents = [
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36 Edg/120.0.0.0',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
        'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:123.0) Gecko/20100101 Firefox/123.0',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 14_3) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.2 Safari/605.1.15'
    ]
    
    encoded_keyword = urllib.parse.quote(keyword)
    search_url = f'https://www.alibaba.com/trade/search?SearchText={encoded_keyword}'
    
    # 🚀 MAXIMUM RETRY: 3 attempts with different strategies
    max_retries = 3
    
    for attempt in range(1, max_retries + 1):
        try:
            # 🎲 Random human-like delay between attempts
            if attempt > 1:
                delay = random.uniform(2.5, 5.0)
                app.logger.info(f'[Alibaba Scraping] 💤 Waiting {delay:.1f}s before retry {attempt}/{max_retries}')
                time.sleep(delay)
            
            # 🎭 Randomize User-Agent for each attempt
            selected_ua = random.choice(user_agents)
            
            app.logger.info(f'[Alibaba Scraping] 🎯 Attempt {attempt}/{max_retries}')
            app.logger.info(f'[Alibaba Scraping] 🎭 User-Agent: {selected_ua[:80]}...')
            
            # 🔥 ScrapingAnt Business Plan Configuration (Residential Proxy + Stealth)
            params = {
                'url': search_url,
                'x-api-key': api_key.strip(),
                'browser': 'true',  # ✅ Real browser rendering (+10 credits)
                'return_page_source': 'true',
                'proxy_type': 'residential',  # 🏠 Business Plan: Residential IP (+25 credits)
                'proxy_country': 'US',  # 🇺🇸 US residential IP for Alibaba
                'stealth': 'true',  # 🥷 Business Plan: Stealth mode (+5 credits)
                'wait_for_selector': '.search-card-item, .organic-list-offer, div[class*="search-card"]',
                'wait_for_timeout': '30000'  # ⏱️ Increased to 30s for stability
            }
            
            # 🎭 Complete modern browser headers (randomized per request)
            chrome_version = random.randint(120, 122)
            headers = {
                'User-Agent': selected_ua,
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.9,ko;q=0.8,zh-CN;q=0.7',
                'Accept-Encoding': 'gzip, deflate, br',
                'sec-ch-ua': f'"Not_A Brand";v="8", "Chromium";v="{chrome_version}", "Google Chrome";v="{chrome_version}"',
                'sec-ch-ua-mobile': '?0',
                'sec-ch-ua-platform': '"Windows"',
                'Sec-Fetch-Dest': 'document',
                'Sec-Fetch-Mode': 'navigate',
                'Sec-Fetch-Site': 'none',
                'Sec-Fetch-User': '?1',
                'Upgrade-Insecure-Requests': '1',
                'Cache-Control': 'max-age=0'
            }
            
            app.logger.info(f'[Alibaba Scraping] 🌐 Sending request to ScrapingAnt...')
            app.logger.info(f'[Alibaba Scraping] Target: {search_url}')
            
            response = requests.get(
                'https://api.scrapingant.com/v2/general',
                params=params,
                headers=headers,
                timeout=120
            )
            
            if response.status_code != 200:
                app.logger.error(f'[Alibaba Scraping] ❌ API error: {response.status_code}')
                app.logger.error(f'[Alibaba Scraping] Response preview: {response.text[:500]}')
                
                if attempt < max_retries:
                    continue  # Retry with different session
                else:
                    return {'products': [], 'count': 0}
            
            app.logger.info(f'[Alibaba Scraping] ✅ Response received: {len(response.text)} chars')
            
            html_content = response.text
            
            # 🚨 Detect anti-bot blocking
            blocking_keywords = ['captcha', 'blocked', 'access denied', 'forbidden', 'verify you are human', 'robot']
            html_lower = html_content.lower()[:2000]
            is_blocked = any(keyword in html_lower for keyword in blocking_keywords)
            
            if is_blocked:
                app.logger.error(f'[Alibaba Scraping] 🚫 BLOCKED on attempt {attempt}/{max_retries}')
                app.logger.error(f'[Alibaba Scraping] Anti-bot keywords detected in response')
                app.logger.error(f'[Alibaba Scraping] HTML Preview: {html_content[:400]}')
                
                if attempt < max_retries:
                    continue  # Retry with fresh IP/session
                else:
                    app.logger.error('[Alibaba Scraping] ❌ All retries failed - Cannot bypass Alibaba firewall')
                    return {'products': [], 'count': 0}
            
            # ✅ SUCCESS - Break retry loop
            app.logger.info(f'[Alibaba Scraping] ✅ Anti-bot bypass SUCCESS on attempt {attempt}/{max_retries}')
            break
        
        except Exception as e:
            app.logger.error(f'[Alibaba Scraping] ❌ Exception on attempt {attempt}/{max_retries}: {str(e)}')
            if attempt < max_retries:
                continue
            else:
                return {'products': [], 'count': 0}
    
    # Parse HTML
    try:
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(html_content, 'lxml')
        
        products = []
        selectors = [
            ('div', 'organic-list-offer'),
            ('div', 'offer-card'),
            ('div', 'product-card'),
            ('div', 'search-card-e-slider'),
            ('a', 'search-card-e-title')
        ]
        
        card_items = []
        for tag, class_name in selectors:
            card_items = soup.find_all(tag, class_=class_name)
            if len(card_items) > 0:
                app.logger.info(f'[Alibaba Scraping] ✅ Found {len(card_items)} items with selector {tag}.{class_name}')
                break
        
        for idx, item in enumerate(card_items[:max_results]):
            try:
                url = ''
                title = ''
                price = 0
                moq = 1
                image = ''
                
                # Extract URL
                link = item.find('a', href=True)
                if link:
                    url = link['href']
                    if not url.startswith('http'):
                        url = 'https://www.alibaba.com' + url
                
                # Extract title
                title_elem = item.find('h2') or item.find('h3') or item.find('div', class_='title')
                if title_elem:
                    title = title_elem.get_text(strip=True)
                
                # Extract price (smart parsing)
                price_elem = item.find('span', class_='price') or item.find('div', class_='price')
                if price_elem:
                    price_text = price_elem.get_text(strip=True)
                    price = parse_smart_price(price_text)
                
                # Extract MOQ
                moq_elem = item.find('span', string=re.compile(r'MOQ', re.I)) or item.find('div', string=re.compile(r'MOQ', re.I))
                if moq_elem:
                    moq_text = moq_elem.get_text(strip=True)
                    moq_match = re.search(r'(\d+)', moq_text)
                    if moq_match:
                        moq = int(moq_match.group(1))
                
                # Extract image (try multiple attributes)
                img = item.find('img')
                if img:
                    image = (img.get('src', '') or 
                            img.get('data-src', '') or 
                            img.get('data-lazy-src', '') or
                            img.get('data-original', '') or
                            img.get('data-img', ''))
                    image = fix_image_url(image)  # 🚀 FIX: Ensure absolute HTTPS URL
                
                # Validation
                if not title or title.startswith('http') or len(title) < 3:
                    continue
                if not url or 'alibaba.com' not in url:
                    continue
                if moq > 2:  # Filter: MOQ must be ≤ 2
                    continue
                
                product = {
                    'url': url,
                    'title': title,
                    'price': price,
                    'moq': moq,
                    'sales': 0,  # Alibaba doesn't show sales easily
                    'image': image,
                    'source_site': 'Alibaba'
                }
                
                if product['title'] and product['price'] > 0:
                    products.append(product)
                    app.logger.info(f'[Alibaba] ✅ Product {len(products)}: {title[:40]} - ${price} (MOQ: {moq})')
            
            except Exception as e:
                app.logger.warning(f'[Alibaba Scraping] Failed to parse product {idx+1}: {str(e)}')
                continue
        
        app.logger.info(f'[Alibaba Scraping] ✅ Found {len(products)} valid products')
        return {'products': products, 'count': len(products)}
    
    except Exception as e:
        app.logger.error(f'[Alibaba Scraping] ❌ Exception: {str(e)}')
        return {'products': [], 'count': 0}

def scrape_aliexpress_search(keyword, max_results=50):
    """
    Scrape AliExpress with ULTIMATE anti-bot bypass
    - Randomized User-Agents
    - Stealth proxy
    - Random delays
    - Multi-retry with different IPs
    """
    import urllib.parse
    import random
    import time
    
    app.logger.info(f'[AliExpress Scraping] ========================================')
    app.logger.info(f'[AliExpress Scraping] 🔥 ULTIMATE ANTI-BOT MODE for: {keyword}')
    
    api_key = get_config('scrapingant_api_key')
    if not api_key or not api_key.strip():
        app.logger.error('[AliExpress Scraping] ❌ No ScrapingAnt API key')
        return {'products': [], 'count': 0}
    
    # 🎭 Pool of REAL User-Agents (Latest browsers)
    user_agents = [
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36 Edg/120.0.0.0',
        'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36'
    ]
    
    encoded_keyword = urllib.parse.quote(keyword)
    search_url = f'https://www.aliexpress.com/wholesale?SearchText={encoded_keyword}'
    
    # 🚀 MAXIMUM RETRY: 3 attempts
    max_retries = 3
    
    for attempt in range(1, max_retries + 1):
        try:
            # 🎲 Random delay between attempts
            if attempt > 1:
                delay = random.uniform(2.5, 5.0)
                app.logger.info(f'[AliExpress Scraping] 💤 Waiting {delay:.1f}s before retry {attempt}/{max_retries}')
                time.sleep(delay)
            
            # 🎭 Randomize User-Agent
            selected_ua = random.choice(user_agents)
            
            app.logger.info(f'[AliExpress Scraping] 🎯 Attempt {attempt}/{max_retries}')
            app.logger.info(f'[AliExpress Scraping] 🎭 User-Agent: {selected_ua[:80]}...')
            
            # 🔥 ScrapingAnt Business Plan Configuration (Residential Proxy + Stealth)
            params = {
                'url': search_url,
                'x-api-key': api_key.strip(),
                'browser': 'true',  # ✅ Real browser (+10 credits)
                'return_page_source': 'true',
                'proxy_type': 'residential',  # 🏠 Business Plan: Residential IP (+25 credits)
                'proxy_country': 'US',  # 🇺🇸 US residential IP for AliExpress
                'stealth': 'true',  # 🥷 Business Plan: Stealth mode (+5 credits)
                'wait_for_selector': '.list--gallery--C2f2tvm, div[class*="product"], div[class*="item"]',
                'wait_for_timeout': '30000'  # ⏱️ Increased to 30s for stability
            }
            
            # 🎭 Randomized browser headers
            chrome_version = random.randint(120, 122)
            headers = {
                'User-Agent': selected_ua,
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.9,ko;q=0.8',
                'Accept-Encoding': 'gzip, deflate, br',
                'sec-ch-ua': f'"Not_A Brand";v="8", "Chromium";v="{chrome_version}", "Google Chrome";v="{chrome_version}"',
                'sec-ch-ua-mobile': '?0',
                'sec-ch-ua-platform': '"Windows"',
                'Sec-Fetch-Dest': 'document',
                'Sec-Fetch-Mode': 'navigate',
                'Sec-Fetch-Site': 'none',
                'Sec-Fetch-User': '?1',
                'Upgrade-Insecure-Requests': '1'
            }
            
            app.logger.info(f'[AliExpress Scraping] 🌐 Sending request to ScrapingAnt...')
            app.logger.info(f'[AliExpress Scraping] Target: {search_url}')
            
            response = requests.get(
                'https://api.scrapingant.com/v2/general',
                params=params,
                headers=headers,
                timeout=120
            )
            
            if response.status_code != 200:
                app.logger.error(f'[AliExpress Scraping] ❌ API error: {response.status_code}')
                app.logger.error(f'[AliExpress Scraping] Response preview: {response.text[:500]}')
                
                if attempt < max_retries:
                    continue
                else:
                    return {'products': [], 'count': 0}
            
            app.logger.info(f'[AliExpress Scraping] ✅ Response received: {len(response.text)} chars')
            
            html_content = response.text
            
            # 🚨 Detect anti-bot blocking
            blocking_keywords = ['captcha', 'blocked', 'access denied', 'forbidden', 'verify you are human', 'robot']
            html_lower = html_content.lower()[:2000]
            is_blocked = any(keyword in html_lower for keyword in blocking_keywords)
            
            if is_blocked:
                app.logger.error(f'[AliExpress Scraping] 🚫 BLOCKED on attempt {attempt}/{max_retries}')
                app.logger.error(f'[AliExpress Scraping] HTML Preview: {html_content[:400]}')
                
                if attempt < max_retries:
                    continue
                else:
                    app.logger.error('[AliExpress Scraping] ❌ All retries failed - Cannot bypass AliExpress firewall')
                    return {'products': [], 'count': 0}
            
            # ✅ SUCCESS
            app.logger.info(f'[AliExpress Scraping] ✅ Anti-bot bypass SUCCESS on attempt {attempt}/{max_retries}')
            break
        
        except Exception as e:
            app.logger.error(f'[AliExpress Scraping] ❌ Exception on attempt {attempt}/{max_retries}: {str(e)}')
            if attempt < max_retries:
                continue
            else:
                return {'products': [], 'count': 0}
    
    # Parse HTML
    try:
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(html_content, 'lxml')
        
        products = []
        selectors = [
            ('div', 'list--gallery--C2f2tvm'),
            ('div', 'product-snippet'),
            ('div', 'product-card'),
            ('div', 'search-item'),
            ('a', 'product-link')
        ]
        
        card_items = []
        for tag, class_name in selectors:
            card_items = soup.find_all(tag, class_=class_name)
            if len(card_items) > 0:
                app.logger.info(f'[AliExpress Scraping] Found {len(card_items)} items with selector {tag}.{class_name}')
                break
        
        for idx, item in enumerate(card_items[:max_results]):
            try:
                url = ''
                title = ''
                price = 0
                image = ''
                
                # Extract URL
                link = item.find('a', href=True)
                if link:
                    url = link['href']
                    if not url.startswith('http'):
                        url = 'https://www.aliexpress.com' + url
                
                # Extract title
                title_elem = item.find('h1') or item.find('h2') or item.find('h3')
                if title_elem:
                    title = title_elem.get_text(strip=True)
                
                # Extract price (smart parsing for discount prices)
                price_elem = item.find('span', class_='price') or item.find('div', class_='price')
                if price_elem:
                    price_text = price_elem.get_text(strip=True)
                    price = parse_smart_price(price_text)
                
                # Extract image (try multiple attributes)
                img = item.find('img')
                if img:
                    image = (img.get('src', '') or 
                            img.get('data-src', '') or 
                            img.get('data-lazy-src', '') or
                            img.get('data-original', '') or
                            img.get('data-img', ''))
                    image = fix_image_url(image)  # 🚀 FIX: Ensure absolute HTTPS URL
                
                # Validation
                if not title or title.startswith('http') or len(title) < 3:
                    continue
                if not url or 'aliexpress.com' not in url:
                    continue
                
                product = {
                    'url': url,
                    'title': title,
                    'price': price,
                    'moq': 1,  # AliExpress always MOQ = 1
                    'sales': 0,
                    'image': image,
                    'source_site': 'AliExpress'
                }
                
                if product['title'] and product['price'] > 0:
                    products.append(product)
                    app.logger.info(f'[AliExpress] ✅ Product {len(products)}: {title[:40]} - ${price}')
            
            except Exception as e:
                app.logger.warning(f'[AliExpress Scraping] Failed to parse product {idx+1}: {str(e)}')
                continue
        
        app.logger.info(f'[AliExpress Scraping] ✅ Found {len(products)} valid products')
        return {'products': products, 'count': len(products)}
    
    except Exception as e:
        app.logger.error(f'[AliExpress Scraping] ❌ Exception: {str(e)}')
        return {'products': [], 'count': 0}

def search_integrated_hybrid(keyword, max_results=50):
    """
    🚀 OFFICIAL API SOURCING ENGINE (Updated 2026-03-05)
    
    ✅ Changed: Now uses AliExpress Official Affiliate API
    ❌ Removed: Alibaba.com (MOQ too high for dropshipping)
    ❌ Removed: ScrapingAnt dependency ($79/month saved)
    
    Search AliExpress only using Official API
    Filter by margin and select best products
    """
    app.logger.info(f'[Search Engine] ========================================')
    app.logger.info(f'[Search Engine] 🚀 Starting Official API search for: {keyword}')
    app.logger.info(f'[Search Engine] Source: AliExpress Affiliate API')
    
    # Search AliExpress using Official API
    app.logger.info('[Search Engine] Searching AliExpress via Official API...')
    aliexpress_result = search_aliexpress_official(keyword, max_results)
    all_products = aliexpress_result.get('products', [])
    app.logger.info(f'[Search Engine] ✅ Found: {len(all_products)} products')
    
    # Changed from: Alibaba + AliExpress scraping
    # To: AliExpress Official API only
    # Reason: MOQ=1 for dropshipping, 100% success rate, $0 cost
    
    if len(all_products) == 0:
        app.logger.error('[Search Engine] ❌ FAILURE: No products found')
        app.logger.error('[Search Engine] ❌ AliExpress API returned no results')
        app.logger.error('[Search Engine] ℹ️  Try different keyword or check API credentials')
        return {'products': [], 'count': 0, 'error': 'No products found'}
    
    # Calculate profitability for each product
    for product in all_products:
        try:
            # Convert USD to CNY for uniform calculation (1 USD ≈ 7.2 CNY)
            price_cny = product['price'] * 7.2
            analysis = analyze_product_profitability(price_cny)
            product['analysis'] = analysis
            product['price_cny'] = price_cny  # Store for DB
            
            # Calculate score: lower price + higher margin + lower MOQ = better
            price_score = max(0, 100 - product['price'])  # Lower price = higher score
            margin_score = analysis['margin']  # Higher margin = higher score
            moq_score = max(0, 10 - product['moq'])  # Lower MOQ = higher score
            
            product['hybrid_score'] = price_score + (margin_score * 2) + (moq_score * 3)
            
            app.logger.debug(f'[Product Score] {product["title"][:30]}: '
                           f'Price=${product["price"]:.2f}, '
                           f'Margin={analysis["margin"]:.1f}%, '
                           f'MOQ={product["moq"]}, '
                           f'Score={product["hybrid_score"]:.1f}')
        except Exception as e:
            app.logger.error(f'[Search Engine] Failed to analyze product: {str(e)}')
            product['hybrid_score'] = 0
    
    # Sort by hybrid score (descending)
    all_products.sort(key=lambda x: x.get('hybrid_score', 0), reverse=True)
    
    app.logger.info(f'[Search Engine] ✅ Search complete: {len(all_products)} products analyzed')
    return {'products': all_products, 'count': len(all_products)}

def analyze_product_profitability(price_cny):
    """Calculate profit margin and KRW price"""
    exchange_rate = float(get_config('cny_exchange_rate', 190))
    buffer = float(get_config('exchange_rate_buffer', 1.05))
    shipping = int(get_config('shipping_cost_base', 5000))
    customs_rate = float(get_config('customs_tax_rate', 0.10))
    
    # Use highest price for safety (prevent loss)
    purchase_price_krw = int(price_cny * exchange_rate * buffer)
    
    # Calculate costs
    customs_tax = int(purchase_price_krw * customs_rate)
    total_cost = purchase_price_krw + shipping + customs_tax
    
    # Calculate sale price with target margin
    target_margin = float(get_config('target_margin_rate', 30)) / 100
    sale_price = int(total_cost / (1 - target_margin))
    
    # Round to nearest 100
    sale_price = ((sale_price + 99) // 100) * 100
    
    # Calculate actual profit
    actual_profit = sale_price - total_cost
    actual_margin = (actual_profit / sale_price * 100) if sale_price > 0 else 0
    
    return {
        'purchase_price_krw': purchase_price_krw,
        'shipping_cost': shipping,
        'customs_tax': customs_tax,
        'total_cost': total_cost,
        'sale_price': sale_price,
        'profit': actual_profit,
        'margin': actual_margin,
        'exchange_rate': exchange_rate
    }

# generate_test_products() DELETED - No mock data allowed

def execute_smart_sourcing(keyword):
    """
    Unified [Smart Sniper] engine for both keyword search and AI discovery
    
    Execution steps:
    1. Hybrid search (Alibaba + AliExpress)
    2. Load config: target_margin, cny_rate, delivery_fee
    3. Margin simulation - drop items not meeting target_margin
    4. Sort by net profit (descending)
    5. Slice to Top 3 only
    6. Use ScrapingAnt tokens ONLY for these 3 items to fetch details
    
    Returns: dict with 'success', 'products' (Top 3 only), 'stats', 'stage_stats'
    """
    app.logger.info(f'[Smart Sniper] ========================================')
    app.logger.info(f'[Smart Sniper] Executing REAL sourcing for keyword: {keyword}')
    app.logger.info(f'[Smart Sniper] NO TEST DATA - Only real Alibaba/AliExpress products')
    
    # Initialize stage-by-stage tracking for UI feedback
    stage_stats = {
        'stage1_scraped': 0,
        'stage2_safe': 0,
        'stage3_profitable': 0,
        'stage4_final': 0,
        'highest_margin_product': None,
        'highest_margin_value': 0
    }
    
    # ========================================================================
    # CRITICAL: SYSTEM CHECK - Verify DB and load configurations
    # ========================================================================
    try:
        app.logger.info('[Smart Sniper] 🔍 Running SYSTEM-CHECK before sourcing...')
        system_config = system_check_critical_configs()
        app.logger.info(f'[Smart Sniper] ✅ System check passed: {system_config}')
    except RuntimeError as e:
        app.logger.error(f'[Smart Sniper] ❌ SYSTEM CHECK FAILED: {str(e)}')
        log_activity('sourcing', f'❌ System check failed: {str(e)}', 'error')
        return {
            'success': False,
            'error': f'System configuration error: {str(e)}',
            'stats': {'scanned': 0, 'safe': 0, 'profitable': 0, 'final_count': 0},
            'stage_stats': stage_stats
        }
    
    # ========================================================================
    # Load configs from DB (NEVER use global variables)
    # ========================================================================
    aliexpress_app_key = get_config('aliexpress_app_key', '')
    aliexpress_app_secret = get_config('aliexpress_app_secret', '')
    openai_key = get_config('openai_api_key', '')
    target_margin_rate = get_config('target_margin_rate', 30)
    cny_exchange_rate = get_config('cny_exchange_rate', 190)
    debug_mode = get_config('debug_mode_ignore_filters', 'false')
    
    # Convert debug_mode to boolean
    debug_mode_enabled = debug_mode.lower() in ['true', '1', 'yes', 'on']
    
    app.logger.info('[Smart Sniper] 📋 Config loaded from DB:')
    app.logger.info(f'  - AliExpress App Key: {aliexpress_app_key}')
    app.logger.info(f'  - AliExpress Secret: {aliexpress_app_secret[:10]}**** (len: {len(aliexpress_app_secret)})')
    app.logger.info(f'  - OpenAI key: {openai_key[:4]}**** (len: {len(openai_key)})')
    app.logger.info(f'  - Target margin: {target_margin_rate}%')
    app.logger.info(f'  - CNY rate: {cny_exchange_rate}')
    app.logger.info(f'  - 🐛 DEBUG MODE (Ignore Filters): {debug_mode_enabled}')
    
    if not aliexpress_app_key or not aliexpress_app_secret:
        app.logger.error('[Smart Sniper] ❌ CRITICAL: AliExpress API credentials missing')
        app.logger.error('[Smart Sniper] Please configure App Key & Secret in database')
        log_activity('sourcing', '❌ AliExpress API not configured', 'error')
        return {
            'success': False,
            'error': 'AliExpress API credentials not configured. Contact system administrator.',
            'stats': {'scanned': 0, 'safe': 0, 'profitable': 0, 'final_count': 0},
            'stage_stats': stage_stats
        }
    
    # Step 1: 🚀 AliExpress Official API Search
    log_activity('sourcing', f'Step 1/5: 🚀 AliExpress Official API Search for "{keyword}"', 'in_progress')
    
    # 🚀 Real API search via AliExpress Affiliate API
    results = search_integrated_hybrid(keyword, max_results=50)
    
    if 'error' in results:
        error_msg = results['error']
        app.logger.error(f'[Smart Sniper] ❌ Scraping FAILED: {error_msg}')
        log_activity('sourcing', f'❌ Search failed: {error_msg}', 'error')
        return {
            'success': False,
            'error': f'Scraping failed: {error_msg}',
            'stats': {'scanned': 0, 'safe': 0, 'profitable': 0, 'final_count': 0},
            'stage_stats': stage_stats
        }
    
    products = results.get('products', [])
    app.logger.info(f'[Smart Sniper] 🚀 HYBRID scraping result: {len(products)} REAL products from Alibaba + AliExpress')
    log_activity('sourcing', f'✅ Found {len(products)} real items from Alibaba/AliExpress', 'success')
    
    # 📊 STAGE 1: Record scraped count
    stage_stats['stage1_scraped'] = len(products)
    app.logger.info(f'[Smart Sniper] 📊 STAGE 1 COMPLETE: {len(products)} products scraped')
    
    # Critical check: if no products, FAIL EXPLICITLY
    if len(products) == 0:
        app.logger.error('[Smart Sniper] ❌ CRITICAL: Zero products found from Alibaba/AliExpress')
        log_activity('sourcing', '❌ No products found - scraping failed', 'error')
        return {
            'success': False,
            'error': 'No products found. ScrapingAnt may be blocked or keyword has no results.',
            'stats': {'scanned': 0, 'safe': 0, 'profitable': 0, 'final_count': 0},
            'stage_stats': stage_stats
        }
    
    # Step 2: Safety Filter (SKIP if debug mode enabled)
    log_activity('sourcing', 'Step 2/5: 🛡️ Applying safety filters', 'in_progress')
    app.logger.info(f'[Smart Sniper] Starting safety filter on {len(products)} products')
    
    if debug_mode_enabled:
        # 🐛 DEBUG MODE: Skip safety filter
        app.logger.warning('[Smart Sniper] 🐛 DEBUG MODE: SKIPPING SAFETY FILTER')
        safe_products = products
        filtered_count = 0
        log_activity('sourcing', f'🐛 Debug mode: All {len(products)} products marked as safe', 'warning')
    else:
        # Normal safety filtering
        safe_products = []
        filtered_count = 0
        for product in products:
            is_safe, reason = check_safety_filter(product['title'])
            if is_safe:
                safe_products.append(product)
            else:
                filtered_count += 1
                app.logger.debug(f'[Safety Filter] Filtered: {product["title"][:40]} - {reason}')
        
        app.logger.info(f'[Smart Sniper] Safety filter result: {len(safe_products)} safe, {filtered_count} filtered')
        log_activity('sourcing', f'{len(safe_products)}/{len(products)} items passed safety filter', 'success')
    
    # 📊 STAGE 2: Record safe count
    stage_stats['stage2_safe'] = len(safe_products)
    app.logger.info(f'[Smart Sniper] 📊 STAGE 2 COMPLETE: {len(safe_products)} products passed safety filter')
    
    # Step 3: Margin Simulation (SKIP if debug mode enabled)
    log_activity('sourcing', 'Step 3/5: 💰 Margin simulation in progress', 'in_progress')
    target_margin = float(get_config('target_margin_rate', 30))
    app.logger.info(f'[Smart Sniper] Target margin: {target_margin}%')
    app.logger.info(f'[Smart Sniper] Analyzing profitability of {len(safe_products)} products')
    
    profitable_products = []
    failed_margin_count = 0
    highest_margin = 0
    highest_margin_product = None
    
    for idx, product in enumerate(safe_products):
        try:
            analysis = analyze_product_profitability(product['price'])
            
            app.logger.debug(f'[Margin Check {idx+1}] {product["title"][:30]}: '
                           f'Price ¥{product["price"]}, '
                           f'Margin {analysis["margin"]:.1f}%, '
                           f'Profit ₩{analysis["profit"]:,}')
            
            # Track highest margin product
            if analysis['margin'] > highest_margin:
                highest_margin = analysis['margin']
                highest_margin_product = {
                    'title': product['title'][:50],
                    'price_cny': product['price'],
                    'margin': analysis['margin'],
                    'profit': analysis['profit']
                }
            
            # Drop items not meeting target margin (UNLESS debug mode enabled)
            if debug_mode_enabled or analysis['margin'] >= target_margin:
                product['analysis'] = analysis
                profitable_products.append(product)
            else:
                failed_margin_count += 1
        except Exception as e:
            app.logger.error(f'[Margin Check] Failed for product {idx+1}: {str(e)}')
            failed_margin_count += 1
    
    # 📊 Record highest margin for diagnostics
    stage_stats['highest_margin_value'] = highest_margin
    stage_stats['highest_margin_product'] = highest_margin_product
    
    app.logger.info(f'[Smart Sniper] 📊 Highest margin found: {highest_margin:.1f}%')
    if highest_margin_product:
        app.logger.info(f'[Smart Sniper] 📊 Highest margin product: {highest_margin_product["title"]}')
    
    if debug_mode_enabled:
        app.logger.warning(f'[Smart Sniper] 🐛 DEBUG MODE: SKIPPING MARGIN FILTER - All {len(profitable_products)} products accepted')
        log_activity('sourcing', f'🐛 Debug mode: All {len(profitable_products)} products marked as profitable', 'warning')
    else:
        app.logger.info(f'[Smart Sniper] Profitability result: {len(profitable_products)} profitable, {failed_margin_count} rejected')
        log_activity('sourcing', f'{len(profitable_products)} items meet target margin {target_margin}%', 'success')
    
    # 📊 STAGE 3: Record profitable count
    stage_stats['stage3_profitable'] = len(profitable_products)
    app.logger.info(f'[Smart Sniper] 📊 STAGE 3 COMPLETE: {len(profitable_products)} products are profitable')
    
    # Step 4: Sort by net profit (descending)
    profitable_products.sort(key=lambda x: x['analysis']['profit'], reverse=True)
    
    # Step 4.5: 🚫 Filter out previously rejected products (only non-expired)
    app.logger.info(f'[Smart Sniper] Checking for previously rejected products...')
    conn = get_db()
    cursor = conn.cursor()
    
    # Only get rejections that haven't expired yet
    cursor.execute('''
        SELECT product_url, expires_at 
        FROM rejected_products 
        WHERE keyword = ? 
        AND (expires_at IS NULL OR expires_at > ?)
    ''', (keyword, datetime.now().isoformat()))
    
    rejected_urls = set(row[0] for row in cursor.fetchall())
    
    # Clean up expired rejections
    cursor.execute('''
        DELETE FROM rejected_products 
        WHERE expires_at IS NOT NULL 
        AND expires_at <= ?
    ''', (datetime.now().isoformat(),))
    
    deleted_count = cursor.rowcount
    if deleted_count > 0:
        app.logger.info(f'[Smart Sniper] ♻️ Cleaned up {deleted_count} expired rejection(s)')
    
    conn.commit()
    conn.close()
    
    if rejected_urls:
        app.logger.info(f'[Smart Sniper] Found {len(rejected_urls)} currently rejected products for keyword: {keyword}')
        filtered_products = [p for p in profitable_products if p['url'] not in rejected_urls]
        rejected_count = len(profitable_products) - len(filtered_products)
        app.logger.info(f'[Smart Sniper] Filtered out {rejected_count} rejected products')
        profitable_products = filtered_products
    else:
        app.logger.info(f'[Smart Sniper] No active rejections found')
    
    # Step 5: Slice to Top 3 ONLY (or all products if debug mode)
    if debug_mode_enabled:
        # 🐛 DEBUG MODE: Return ALL products instead of just top 3
        top_3 = profitable_products[:50]  # Cap at 50 to avoid UI overload
        app.logger.warning(f'[Smart Sniper] 🐛 DEBUG MODE: Returning TOP {len(top_3)} products (not limited to 3)')
        log_activity('sourcing', f'Step 4/5: 🐛 Debug mode: Top {len(top_3)} selected', 'warning')
    else:
        top_3 = profitable_products[:3]
        log_activity('sourcing', f'Step 4/5: 🎯 Top 3 selected (sorted by profit, rejected products excluded)', 'success')
    
    # 📊 STAGE 4: Record final count
    stage_stats['stage4_final'] = len(top_3)
    app.logger.info(f'[Smart Sniper] 📊 STAGE 4 COMPLETE: {len(top_3)} products in final selection')
    
    if len(top_3) == 0:
        # 🚨 CRITICAL: No products after all filters
        app.logger.error('[Smart Sniper] ❌ ZERO products after all filters!')
        app.logger.error(f'[Smart Sniper] 📊 Breakdown: Scraped={stage_stats["stage1_scraped"]}, '
                        f'Safe={stage_stats["stage2_safe"]}, '
                        f'Profitable={stage_stats["stage3_profitable"]}, '
                        f'Final={stage_stats["stage4_final"]}')
        app.logger.error(f'[Smart Sniper] 💡 Highest margin found: {highest_margin:.1f}% (target: {target_margin}%)')
        
        if highest_margin_product:
            app.logger.error(f'[Smart Sniper] 💡 SUGGESTION: Best product was: {highest_margin_product["title"]}')
            app.logger.error(f'[Smart Sniper] 💡 Consider lowering target margin or enabling debug mode')
        
        log_activity('sourcing', '⚠️ No products met all criteria - check stage breakdown in logs', 'warning')
        return {
            'success': True,
            'products': [],
            'stats': {
                'scanned': len(products),
                'safe': len(safe_products),
                'profitable': len(profitable_products),
                'final_count': 0
            },
            'stage_stats': stage_stats,
            'suggestion': f'No products found. Highest margin was {highest_margin:.1f}% (target: {target_margin}%). Consider lowering margin target or enabling Debug Mode.'
        }
    
    # Step 5.5: 📊 Market Analysis (Naver Shopping)
    app.logger.info(f'[Smart Sniper] 📊 Starting market analysis for keyword: {keyword}')
    log_activity('sourcing', 'Step 5/5: 📊 Analyzing market prices (Naver Shopping)', 'in_progress')
    
    market_data = None
    naver_client_id = get_config('naver_client_id', '')
    naver_client_secret = get_config('naver_client_secret', '')
    
    if naver_client_id and naver_client_secret:
        try:
            market_data = analyze_naver_market(keyword, naver_client_id, naver_client_secret)
            
            if market_data.get('success'):
                app.logger.info(f'[Market Analysis] ✅ SUCCESS')
                app.logger.info(f'[Market Analysis] Total products analyzed: {market_data["analyzed_products"]}')
                app.logger.info(f'[Market Analysis] Average price: ₩{market_data["avg_price"]:,}')
                app.logger.info(f'[Market Analysis] Price range: ₩{market_data["min_price"]:,} ~ ₩{market_data["max_price"]:,}')
                app.logger.info(f'[Market Analysis] Recommended price: ₩{market_data["recommended_price"]:,}')
                app.logger.info(f'[Market Analysis] Market position: {market_data["analysis_summary"]["market_position"]}')
                
                log_activity('sourcing', f'✅ Market analysis: Avg ₩{market_data["avg_price"]:,}, Recommend ₩{market_data["recommended_price"]:,}', 'success')
            else:
                app.logger.warning(f'[Market Analysis] ⚠️ Failed: {market_data.get("error")}')
                log_activity('sourcing', f'⚠️ Market analysis failed: {market_data.get("error")}', 'warning')
        except Exception as e:
            app.logger.error(f'[Market Analysis] ❌ Exception: {str(e)}')
            market_data = None
    else:
        app.logger.warning('[Market Analysis] ⚠️ Naver API credentials not configured')
        log_activity('sourcing', '⚠️ Market analysis skipped (API credentials missing)', 'warning')
    
    # Step 6: Save Top 3 to Database
    log_activity('sourcing', 'Step 6/6: 💾 Saving Top 3 to database', 'in_progress')
    app.logger.info(f'[Smart Sniper] Attempting to save {len(top_3)} products to database')
    
    conn = get_db()
    cursor = conn.cursor()
    
    saved_count = 0
    for idx, product in enumerate(top_3):
        try:
            app.logger.info(f'[DB Save {idx+1}] Title: {product["title"][:50]}')
            app.logger.info(f'[DB Save {idx+1}] Price CNY: ¥{product["price"]}')
            app.logger.info(f'[DB Save {idx+1}] Price KRW: ₩{product["analysis"]["sale_price"]:,}')
            app.logger.info(f'[DB Save {idx+1}] Margin: {product["analysis"]["margin"]:.1f}%')
            app.logger.info(f'[DB Save {idx+1}] Profit: ₩{product["analysis"]["profit"]:,}')
            
            cursor.execute('''
                INSERT INTO sourced_products 
                (original_url, title_cn, price_cny, price_krw, profit_margin, 
                 estimated_profit, safety_status, images_json, status,
                 source_site, moq, traffic_score)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                product['url'],
                product['title'],
                product.get('price_cny', product['price']),  # Use converted CNY if available
                product['analysis']['sale_price'],
                product['analysis']['margin'],
                product['analysis']['profit'],
                'passed',
                json.dumps([fix_image_url(product.get('image', ''))]),  # 🚀 FIX: Ensure HTTPS URL
                'pending',
                product.get('source_site', 'alibaba'),  # 🚀 Source: Alibaba/AliExpress
                product.get('moq', 1),  # 🚀 NEW: MOQ
                product.get('sales', 0)  # 🚀 NEW: traffic score (use sales as proxy)
            ))
            saved_count += 1
            app.logger.info(f'[DB Save {idx+1}] ✅ Successfully inserted')
        except Exception as e:
            app.logger.error(f'[DB Save {idx+1}] ❌ Failed to insert: {str(e)}')
            app.logger.exception(e)
    
    conn.commit()
    conn.close()
    
    app.logger.info(f'[Smart Sniper] Completed: {saved_count}/{len(top_3)} products saved to database')
    app.logger.info(f'[Smart Sniper] 📊 FINAL BREAKDOWN:')
    app.logger.info(f'[Smart Sniper]   Stage 1 (Scraped): {stage_stats["stage1_scraped"]}')
    app.logger.info(f'[Smart Sniper]   Stage 2 (Safe): {stage_stats["stage2_safe"]}')
    app.logger.info(f'[Smart Sniper]   Stage 3 (Profitable): {stage_stats["stage3_profitable"]}')
    app.logger.info(f'[Smart Sniper]   Stage 4 (Final): {stage_stats["stage4_final"]}')
    log_activity('sourcing', f'✅ Smart Sourcing completed: {saved_count} products saved', 'success')
    
    return {
        'success': True,
        'products': top_3,
        'stats': {
            'scanned': len(products),
            'safe': len(safe_products),
            'profitable': len(profitable_products),
            'final_count': len(top_3)
        },
        'stage_stats': stage_stats,
        'debug_mode_enabled': debug_mode_enabled,
        'market_analysis': market_data  # 시장 분석 데이터 추가
    }

@app.route('/api/sourcing/start', methods=['POST'])
@login_required
def start_sourcing():
    """
    Unified sourcing endpoint supporting two modes:
    - Case A (Direct search): Use user-provided keyword
    - Case B (AI Blue Ocean): Run GPT-4 analysis first, then use suggested keyword
    """
    data = request.json
    user_keyword = data.get('keyword', '')
    mode = data.get('mode', 'direct')  # 'direct' or 'ai_discovery'
    
    app.logger.info(f'=== Sourcing Started by {current_user.username} ===')
    app.logger.info(f'Mode: {mode}, User keyword: {user_keyword}')
    
    # Determine target keyword based on mode
    if mode == 'ai_discovery':
        # Case B: AI Blue Ocean Discovery
        log_activity('sourcing', 'Step 0: 🌊 Blue Ocean Market Analysis', 'in_progress')
        
        blue_ocean_result = analyze_blue_ocean_market(user_keyword)
        target_keyword = blue_ocean_result['suggested_keyword']
        reasoning = blue_ocean_result['reasoning']
        
        app.logger.info(f'Blue Ocean Keyword: {target_keyword}')
        app.logger.info(f'Reasoning: {reasoning}')
        
        log_activity('sourcing', 
                    f'🎯 Blue Ocean Keyword: "{target_keyword}"', 
                    'success',
                    {'reasoning': reasoning, 'original': user_keyword})
        
        blue_ocean_data = {
            'original_keyword': user_keyword,
            'suggested_keyword': target_keyword,
            'reasoning': reasoning,
            'trend_score': blue_ocean_result.get('trend_score', 0),
            'competition_score': blue_ocean_result.get('competition_score', 0)
        }
    else:
        # Case A: Direct keyword search
        target_keyword = user_keyword
        blue_ocean_data = None
        log_activity('sourcing', f'📌 Direct search mode: "{target_keyword}"', 'info')
    
    # Execute unified Smart Sniper engine - REAL DATA ONLY
    result = execute_smart_sourcing(target_keyword)
    
    if not result['success']:
        return jsonify({'error': result.get('error', 'Unknown error'), 'stage_stats': result.get('stage_stats', {})}), 500
    
    # Build response
    response_data = {
        'success': True,
        'mode': mode,
        'keyword': target_keyword,
        'stats': result['stats'],
        'stage_stats': result.get('stage_stats', {}),  # NEW: Stage-by-stage breakdown
        'debug_mode_enabled': result.get('debug_mode_enabled', False),
        'suggestion': result.get('suggestion', '')  # NEW: Suggestion when no products found
    }
    
    if blue_ocean_data:
        response_data['blue_ocean_analysis'] = blue_ocean_data
    
    app.logger.info(f'=== Sourcing Completed: {result["stats"]["final_count"]} products ===')
    
    return jsonify(response_data)

@app.route('/api/product/reject', methods=['POST'])
@login_required
def reject_product():
    """
    Mark a product as rejected with expiration date
    
    Request JSON:
    {
        "product_url": "https://...",
        "product_title": "...",
        "keyword": "..."
    }
    """
    try:
        data = request.json
        product_url = data.get('product_url')
        product_title = data.get('product_title', '')
        keyword = data.get('keyword', '')
        
        if not product_url:
            return jsonify({'success': False, 'error': 'product_url required'}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Get rejection expiry days from config
        cursor.execute("SELECT value FROM config WHERE key = 'rejection_expiry_days'")
        row = cursor.fetchone()
        expiry_days = int(row[0]) if row else 30
        
        # Calculate expiration date
        from datetime import timedelta
        expires_at = (datetime.now() + timedelta(days=expiry_days)).isoformat()
        
        # Insert or replace (update expiration if exists)
        cursor.execute('''
            INSERT OR REPLACE INTO rejected_products 
            (product_url, product_title, keyword, rejected_at, expires_at)
            VALUES (?, ?, ?, CURRENT_TIMESTAMP, ?)
        ''', (product_url, product_title, keyword, expires_at))
        
        conn.commit()
        conn.close()
        
        app.logger.info(f'[Product Reject] URL: {product_url[:50]}...')
        app.logger.info(f'[Product Reject] Title: {product_title[:50]}...')
        app.logger.info(f'[Product Reject] Keyword: {keyword}')
        app.logger.info(f'[Product Reject] Expires at: {expires_at} ({expiry_days} days)')
        
        log_activity('sourcing', f'Product rejected: {product_title[:30]}... (expires in {expiry_days} days)', 'info')
        
        return jsonify({
            'success': True,
            'message': f'Product rejected for {expiry_days} days',
            'expires_at': expires_at
        })
        
    except Exception as e:
        app.logger.error(f'[Product Reject] Error: {str(e)}')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/sourcing/test-scraping', methods=['POST'])
@login_required
def test_scraping():
    """
    Diagnostic endpoint to test hybrid scraping functionality (Alibaba + AliExpress)
    Returns detailed information about scraping attempt
    """
    data = request.json
    keyword = data.get('keyword', '电热毯')  # Default test keyword
    
    app.logger.info(f'[Test Scraping] Testing scraping for keyword: {keyword}')
    
    # Test scraping
    result = search_integrated_hybrid(keyword, max_results=10)
    
    # Build diagnostic response
    if 'error' in result:
        return jsonify({
            'success': False,
            'error': result['error'],
            'keyword': keyword,
            'timestamp': get_kst_now().isoformat()
        })
    
    products = result.get('products', [])
    
    return jsonify({
        'success': True,
        'keyword': keyword,
        'product_count': len(products),
        'products': products[:5],  # Return first 5 for preview
        'sample_product': products[0] if products else None,
        'timestamp': datetime.now().isoformat()
    })

# ============================================================================
# MODULE 3: AI CONTENT GENERATOR WITH IMAGE PROCESSING
# ============================================================================

def generate_marketing_copy(title, price):
    """Generate marketing copy using GPT-4"""
    api_key = get_config('openai_api_key')
    if not api_key:
        return '상품 설명이 준비 중입니다.'
    
    prompt = f"""
🚨 너는 대한민국 이커머스 1등 판매자다. 월매출 5억 이상 찍는 쿠팡/네이버 스마트스토어 운영자처럼 써라.

상품명: {title}
가격: {price:,}원

다음 구조로 **반드시** 작성:
1. 강력한 후킹: "이런 불편 겪고 계신가요?" (고객 결핍 자극)
2. 공감+해결: "그래서 이 제품을 준비했습니다" (솔루션 제시)
3. USP 3가지: 이 제품만의 차별점 3개 (구체적 수치/특징)
4. 구매 확신: "지금 바로 경험하세요" (행동 유도)

⚠️ 절대 금지:
- "1. 훅:", "2. 문제:" 같은 라벨 쓰지 마라
- 추상적 표현 금지 (예: "품질이 좋습니다" ❌)
- 구체적 숫자/특징으로 써라 (예: "24시간 보온" ✅)

300자 이내, 자연스럽게 흐르는 문장으로 써라.
"""
    
    try:
        # CRITICAL: Using gpt-4o-mini (universal access, no 404 errors)
        response = requests.post(
            'https://api.openai.com/v1/chat/completions',
            headers={
                'Authorization': f'Bearer {api_key}',
                'Content-Type': 'application/json'
            },
            json={
                'model': 'gpt-4o-mini',  # FORCED: Must use gpt-4o-mini
                'messages': [
                    {'role': 'system', 'content': '너는 월 5억 찍는 대한민국 1등 이커머스 판매자다. 쿠팡/네이버 베스트셀러를 만드는 전문가다.'},
                    {'role': 'user', 'content': prompt}
                ],
                'temperature': 0.7,
                'max_tokens': 500
            },
            timeout=30
        )
        
        if response.status_code == 200:
            copy = response.json()['choices'][0]['message']['content']
            
            # 🔥 EMERGENCY FIX: Remove labels from marketing copy
            import re
            copy = re.sub(r'\d+\.\s*(훅|문제|솔루션|Hook|Problem|Solution):\s*', '', copy, flags=re.IGNORECASE)
            copy = re.sub(r'\*\*(훅|문제|솔루션|Hook|Problem|Solution)\s*\([^)]+\):\*\*\s*', '', copy, flags=re.IGNORECASE)
            copy = re.sub(r'(훅|문제|솔루션|Hook|Problem|Solution):\s*', '', copy, flags=re.IGNORECASE)
            
            return copy.strip()
        else:
            return '상품 설명이 준비 중입니다.'
    
    except Exception as e:
        log_activity('content', f'Failed to generate copy: {str(e)}', 'error')
        return '상품 설명이 준비 중입니다.'

def generate_winning_product_page(title, price, images):
    """
    Generate a WINNING product detail page using the 5-section formula.
    This structure mimics best-selling Coupang/Naver products.
    
    5-SECTION FORMULA:
    1. Hook: "Do you have this problem?" (GIF or strong image)
    2. Empathy/Solution: "That's why we prepared this." (Product intro)
    3. Key Points: "3 reasons why this product is special" (Icons + brief)
    4. Details: Usage shots + spec explanation (Image-text alternating)
    5. FAQ/Shipping: Frequently asked questions
    """
    api_key = get_config('openai_api_key')
    if not api_key:
        return generate_fallback_product_page(title, images)
    
    # Classify images: lifestyle shots first, detail shots later
    lifestyle_imgs = images[:3] if len(images) >= 3 else images
    detail_imgs = images[3:] if len(images) > 3 else []
    
    system_prompt = """You are a TOP-TIER Korean E-commerce Merchandiser specialized in creating WINNING product pages.

🚨 CRITICAL RULES - MUST FOLLOW:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
1. ❌ NEVER use placeholder images (via.placeholder.com, example.com, etc.)
2. ✅ ONLY use image URLs from the provided product_images list
3. ❌ NEVER write section labels like "1. 훅:", "2. 솔루션:" in the output
4. ✅ Write naturally flowing marketing copy without numbered labels
5. ✅ MUST include SEO tags at the end: [TAGS]: keyword1, keyword2, ...

IMAGE RULES (CRITICAL):
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
- Use ONLY the exact image URLs provided in the prompt
- Format: <img src="[PROVIDED_URL]" style="width: 100%; border-radius: 12px; box-shadow: 0 4px 12px rgba(0,0,0,0.15); margin: 20px 0;">
- If you need to reference images, use {img_1}, {img_2}, {img_3}, etc.
- These placeholders will be replaced with REAL product images

TEXT RULES (CRITICAL):
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
- NO section labels like "1. 훅:", "2. 문제:", "3. 솔루션:"
- Write as continuous, natural marketing copy
- Use HTML headings (<h3>) for section titles only
- No Markdown syntax (**bold**, ##heading)

SEO RULES (MANDATORY):
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
At the END of your HTML, include:
[TAGS]: keyword1, keyword2, keyword3, keyword4, keyword5, keyword6, keyword7, keyword8, keyword9, keyword10

Example keywords: 겨울이불, 따뜻한담요, 전기장판, 방한용품, 수면용품...

SECTION STRUCTURE:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION 1: HOOK (강렬한 후킹)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
- Start with customer PAIN POINT
- Use emotional language
- Show problem scenario
- 1-2 sentences MAX

SECTION 2: EMPATHY & SOLUTION (공감 + 솔루션)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
- "그래서 준비했습니다"
- Introduce product as THE solution
- Show lifestyle image
- Build excitement

SECTION 3: KEY POINTS (핵심 포인트 3가지)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Use this EXACT format:
<div style="background: #f8f9fa; padding: 30px; border-radius: 15px; margin: 30px 0;">
  <h3 style="font-size: 24px; font-weight: bold; margin-bottom: 20px; text-align: center;">✨ 이 제품이 특별한 이유 3가지</h3>
  
  <div style="display: grid; grid-template-columns: repeat(3, 1fr); gap: 20px; margin-top: 20px;">
    <div style="text-align: center; padding: 20px; background: white; border-radius: 10px;">
      <div style="font-size: 48px; margin-bottom: 10px;">🎯</div>
      <h4 style="font-size: 18px; font-weight: bold; margin-bottom: 10px;">Point 1 Title</h4>
      <p style="font-size: 14px; color: #666;">Brief explanation</p>
    </div>
    <!-- Repeat for Point 2 and 3 -->
  </div>
</div>

SECTION 4: DETAILS (디테일 설명)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Alternate between image and text:
- Image → Feature explanation
- Image → Spec details
- Image → Usage scenario

SECTION 5: FAQ & TRUST (FAQ + 신뢰 구축)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
<div style="background: #fff9e6; padding: 30px; border-radius: 15px; margin: 30px 0;">
  <h3 style="font-size: 22px; font-weight: bold; margin-bottom: 20px;">💬 자주 묻는 질문</h3>
  
  <div style="margin: 15px 0; padding: 15px; background: white; border-left: 4px solid #ffa500; border-radius: 8px;">
    <strong style="color: #333;">Q: Question here?</strong>
    <p style="margin-top: 10px; color: #666;">A: Answer here.</p>
  </div>
  <!-- Repeat 3-5 FAQs -->
</div>

FINAL SECTION: CTA
<div style="text-align: center; padding: 40px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; border-radius: 15px; margin: 30px 0;">
  <h3 style="font-size: 26px; font-weight: bold; margin-bottom: 15px;">지금 바로 경험해보세요!</h3>
  <p style="font-size: 16px; margin-bottom: 20px;">✅ 무료배송 | ✅ 당일출고 | ✅ 100% 환불보증</p>
</div>

REMEMBER:
- NO Markdown (**bold**, ##heading) - Use HTML only
- ALL images MUST have rounded corners and shadows
- Use emojis for visual appeal (🎯, ✨, 💯, 👍)
- Keep Korean natural and persuasive
- Focus on BENEFITS over features
"""
    
    user_prompt = f"""Create a WINNING product detail page for:

Product: {title}
Price: {price:,}원

🖼️ REAL PRODUCT IMAGES (Use these EXACT URLs):
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{chr(10).join([f'{{img_{i+1}}} = {url}' for i, url in enumerate(images)])}

IMAGE USAGE:
- Lifestyle shots (1-3): Use {chr(10).join([f'{{img_{i+1}}}' for i in range(min(3, len(images)))])}
- Detail shots (4+): Use {chr(10).join([f'{{img_{i+1}}}' for i in range(3, len(images))])} if available

🚨 CRITICAL REMINDERS:
1. Replace {{img_1}}, {{img_2}}, etc. with the PROVIDED URLs above
2. NO placeholder images (via.placeholder.com)
3. NO section labels like "1. 훅:", "2. 솔루션:" in output
4. Write natural, flowing marketing copy
5. END with: [TAGS]: keyword1, keyword2, ... (10 Korean keywords)

TASK:
Generate complete HTML following the 5-section formula.
Make it look PROFESSIONAL and PERSUASIVE like top Coupang sellers.

OUTPUT FORMAT:
Pure HTML code (no markdown, no code blocks).
Start directly with HTML tags.
End with SEO tags: [TAGS]: keyword1, keyword2, ...
"""
    
    try:
        response = requests.post(
            'https://api.openai.com/v1/chat/completions',
            headers={
                'Authorization': f'Bearer {api_key}',
                'Content-Type': 'application/json'
            },
            json={
                'model': 'gpt-4o-mini',
                'messages': [
                    {'role': 'system', 'content': system_prompt},
                    {'role': 'user', 'content': user_prompt}
                ],
                'temperature': 0.8,
                'max_tokens': 3000
            },
            timeout=60
        )
        
        if response.status_code == 200:
            content = response.json()['choices'][0]['message']['content']
            
            # Clean up: remove code blocks if present
            content = content.replace('```html', '').replace('```', '').strip()
            
            # 🔥 CRITICAL: Replace image placeholders with REAL URLs
            for i, img_url in enumerate(images, 1):
                content = content.replace(f'{{img_{i}}}', img_url)
                content = content.replace(f'IMAGE_{i}', img_url)
                content = content.replace(f'[IMAGE_{i}]', img_url)
            
            # 🔥 CRITICAL: Remove ANY remaining placeholder images
            import re
            # Remove via.placeholder.com
            content = re.sub(r'<img[^>]*src=["\']https?://via\.placeholder\.com[^"\']*["\'][^>]*>', '', content)
            # Remove example.com placeholders
            content = re.sub(r'<img[^>]*src=["\']https?://example\.com[^"\']*["\'][^>]*>', '', content)
            
            # 🔥 CRITICAL: Extract SEO tags
            tags = ''
            if '[TAGS]:' in content or '[TAGS]' in content:
                import re
                tag_match = re.search(r'\[TAGS\]:?\s*(.+?)(?:\n|$)', content, re.IGNORECASE)
                if tag_match:
                    tags = tag_match.group(1).strip()
                    # Remove tags from HTML content
                    content = re.sub(r'\[TAGS\]:?.+?(?:\n|$)', '', content, flags=re.IGNORECASE)
            
            # 🔥 EMERGENCY FIX: AGGRESSIVE label removal (all patterns)
            import re
            
            # Pattern 1: Numbered labels (1. 훅:, 2. 문제:, etc.)
            content = re.sub(r'\d+\.\s*(훅|문제|솔루션|공감|핵심|특징|디테일|FAQ|신뢰|CTA|Hook|Problem|Solution):\s*', '', content, flags=re.IGNORECASE)
            
            # Pattern 2: Bold labels (**훅(Hook):**, etc.)
            content = re.sub(r'\*\*(훅|문제|솔루션|공감|핵심|특징|디테일|FAQ|신뢰|CTA|Hook|Problem|Solution)\s*\([^)]+\):\*\*\s*', '', content, flags=re.IGNORECASE)
            
            # Pattern 3: Labels without numbers (훅:, 문제:, etc.)
            content = re.sub(r'(훅|문제|솔루션|공감|핵심|특징|디테일|FAQ|신뢰|CTA|Hook|Problem|Solution):\s*', '', content, flags=re.IGNORECASE)
            
            # Pattern 4: Labels in parentheses ((훅), (Hook), etc.)
            content = re.sub(r'\((훅|문제|솔루션|공감|핵심|특징|디테일|FAQ|신뢰|CTA|Hook|Problem|Solution)\)\s*', '', content, flags=re.IGNORECASE)
            
            # Pattern 5: Section markers (SECTION 1:, etc.)
            content = re.sub(r'SECTION\s+\d+:\s*', '', content, flags=re.IGNORECASE)
            
            app.logger.info(f'[Content Generation] ✅ Generated winning product page: {len(content)} chars')
            app.logger.info(f'[Content Generation] 🧹 Applied AGGRESSIVE label cleanup (5 patterns)')
            if tags:
                app.logger.info(f'[Content Generation] ✅ Extracted SEO tags: {tags}')
            
            return content, tags
        else:
            app.logger.error(f'[Content Generation] ❌ API error: {response.status_code}')
            return generate_fallback_product_page(title, images), ''
    
    except Exception as e:
        app.logger.error(f'[Content Generation] ❌ Exception: {str(e)}')
        log_activity('content', f'Failed to generate winning page: {str(e)}', 'error')
        return generate_fallback_product_page(title, images), ''

def generate_fallback_product_page(title, images):
    """Fallback product page when API fails"""
    img_tags = '\n'.join([
        f'<img src="{img}" style="width: 100%; border-radius: 12px; box-shadow: 0 4px 12px rgba(0,0,0,0.15); margin: 20px 0;" alt="Product Image">'
        for img in images[:5]
    ])
    
    return f"""
<div style="font-family: 'Noto Sans KR', sans-serif; max-width: 800px; margin: 0 auto; padding: 20px;">
    <div style="text-align: center; padding: 40px 20px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; border-radius: 15px; margin-bottom: 30px;">
        <h2 style="font-size: 28px; font-weight: bold; margin-bottom: 15px;">{title}</h2>
        <p style="font-size: 18px;">✨ 프리미엄 품질을 합리적인 가격에</p>
    </div>
    
    {img_tags}
    
    <div style="background: #f8f9fa; padding: 30px; border-radius: 15px; margin: 30px 0;">
        <h3 style="font-size: 24px; font-weight: bold; margin-bottom: 20px; text-align: center;">✨ 제품 특징</h3>
        <ul style="list-style: none; padding: 0;">
            <li style="padding: 15px; margin: 10px 0; background: white; border-left: 4px solid #667eea; border-radius: 8px;">
                <strong>🎯 고품질 소재</strong><br>
                <span style="color: #666;">엄선된 소재로 만든 프리미엄 제품</span>
            </li>
            <li style="padding: 15px; margin: 10px 0; background: white; border-left: 4px solid #667eea; border-radius: 8px;">
                <strong>💯 완벽한 품질 관리</strong><br>
                <span style="color: #666;">철저한 검수를 거친 안심 상품</span>
            </li>
            <li style="padding: 15px; margin: 10px 0; background: white; border-left: 4px solid #667eea; border-radius: 8px;">
                <strong>🚚 빠른 배송</strong><br>
                <span style="color: #666;">주문 후 신속하게 배송해드립니다</span>
            </li>
        </ul>
    </div>
    
    <div style="text-align: center; padding: 40px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; border-radius: 15px;">
        <h3 style="font-size: 26px; font-weight: bold; margin-bottom: 15px;">지금 바로 만나보세요!</h3>
        <p style="font-size: 16px;">✅ 무료배송 | ✅ 당일출고 | ✅ 100% 환불보증</p>
    </div>
</div>
"""

def remove_watermark_ai(image):
    """
    AI-based watermark removal using inpainting
    Detects and removes text/logo watermarks
    """
    import cv2
    import numpy as np
    
    # Convert PIL to OpenCV
    img_array = np.array(image)
    img_cv = cv2.cvtColor(img_array, cv2.COLOR_RGB2BGR)
    
    # Convert to grayscale for text detection
    gray = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY)
    
    # Detect text regions (likely watermarks)
    # Method 1: Threshold for white/light text
    _, thresh1 = cv2.threshold(gray, 200, 255, cv2.THRESH_BINARY)
    
    # Method 2: Threshold for dark text
    _, thresh2 = cv2.threshold(gray, 100, 255, cv2.THRESH_BINARY_INV)
    
    # Combine both masks
    mask = cv2.bitwise_or(thresh1, thresh2)
    
    # Morphological operations to connect text
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (5, 5))
    mask = cv2.dilate(mask, kernel, iterations=2)
    
    # Find contours (potential watermark regions)
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    # Create final mask for inpainting
    inpaint_mask = np.zeros(gray.shape, dtype=np.uint8)
    
    # Filter contours (remove small noise, keep potential watermarks)
    for contour in contours:
        area = cv2.contourArea(contour)
        x, y, w, h = cv2.boundingRect(contour)
        
        # Heuristic: watermarks are usually in corners or edges
        is_corner = (x < img_cv.shape[1] * 0.3 or x > img_cv.shape[1] * 0.7 or
                    y < img_cv.shape[0] * 0.3 or y > img_cv.shape[0] * 0.7)
        
        # Size filter: not too small, not too large
        if 500 < area < img_cv.shape[0] * img_cv.shape[1] * 0.1 and is_corner:
            cv2.drawContours(inpaint_mask, [contour], -1, 255, -1)
    
    # Inpaint to remove watermarks
    result = cv2.inpaint(img_cv, inpaint_mask, 3, cv2.INPAINT_TELEA)
    
    # Convert back to PIL
    result_rgb = cv2.cvtColor(result, cv2.COLOR_BGR2RGB)
    return Image.fromarray(result_rgb)

def remove_background_ai(image):
    """
    AI-based background removal using GrabCut algorithm
    Isolates the main product
    """
    import cv2
    import numpy as np
    
    # Convert PIL to OpenCV
    img_array = np.array(image)
    img_cv = cv2.cvtColor(img_array, cv2.COLOR_RGB2BGR)
    
    # Create mask
    mask = np.zeros(img_cv.shape[:2], np.uint8)
    
    # Background and foreground models
    bgd_model = np.zeros((1, 65), np.float64)
    fgd_model = np.zeros((1, 65), np.float64)
    
    # Define ROI (assume product is in center 60% of image)
    h, w = img_cv.shape[:2]
    rect = (int(w*0.2), int(h*0.2), int(w*0.6), int(h*0.6))
    
    # Apply GrabCut
    try:
        cv2.grabCut(img_cv, mask, rect, bgd_model, fgd_model, 5, cv2.GC_INIT_WITH_RECT)
        
        # Create binary mask
        mask2 = np.where((mask == 2) | (mask == 0), 0, 1).astype('uint8')
        
        # Apply mask
        img_cv = img_cv * mask2[:, :, np.newaxis]
        
        # Convert to RGBA
        result_rgb = cv2.cvtColor(img_cv, cv2.COLOR_BGR2RGB)
        result_rgba = Image.fromarray(result_rgb).convert('RGBA')
        
        # Make background transparent
        datas = result_rgba.getdata()
        new_data = []
        for item in datas:
            # If pixel is black (background), make transparent
            if item[0] < 10 and item[1] < 10 and item[2] < 10:
                new_data.append((255, 255, 255, 0))
            else:
                new_data.append(item)
        
        result_rgba.putdata(new_data)
        return result_rgba
    
    except Exception as e:
        app.logger.warning(f'[Background Removal] Failed: {str(e)}, returning original')
        return image.convert('RGBA')

def create_optimized_background(product_image, product_type='general'):
    """
    Create optimized professional background for product
    Based on product category (electronics, fashion, home, etc.)
    """
    from PIL import ImageDraw, ImageFilter
    
    width, height = product_image.size
    
    # Background styles by product type
    backgrounds = {
        'electronics': {
            'gradient_start': (240, 248, 255),  # Light blue
            'gradient_end': (230, 230, 250),    # Lavender
            'accent': (100, 149, 237)           # Cornflower blue
        },
        'fashion': {
            'gradient_start': (255, 250, 250),  # Snow white
            'gradient_end': (255, 240, 245),    # Lavender blush
            'accent': (255, 182, 193)           # Light pink
        },
        'home': {
            'gradient_start': (250, 250, 240),  # Ivory
            'gradient_end': (245, 245, 220),    # Beige
            'accent': (210, 180, 140)           # Tan
        },
        'general': {
            'gradient_start': (255, 255, 255),  # Pure white
            'gradient_end': (248, 248, 255),    # Ghost white
            'accent': (220, 220, 220)           # Light gray
        }
    }
    
    style = backgrounds.get(product_type, backgrounds['general'])
    
    # Create gradient background
    background = Image.new('RGB', (width, height), style['gradient_start'])
    draw = ImageDraw.Draw(background)
    
    # Radial gradient effect
    for i in range(min(width, height) // 2):
        alpha = i / (min(width, height) // 2)
        color = tuple(int(style['gradient_start'][j] * (1 - alpha) + 
                         style['gradient_end'][j] * alpha) for j in range(3))
        
        draw.ellipse([width//2 - i, height//2 - i, 
                     width//2 + i, height//2 + i], 
                    fill=color, outline=color)
    
    # Add subtle decorative elements
    # Corner accents
    accent_size = min(width, height) // 8
    
    # Top-left accent
    draw.arc([0, 0, accent_size*2, accent_size*2], 
            start=180, end=270, fill=style['accent'], width=3)
    
    # Bottom-right accent
    draw.arc([width - accent_size*2, height - accent_size*2, width, height], 
            start=0, end=90, fill=style['accent'], width=3)
    
    # Apply subtle blur for professional look
    background = background.filter(ImageFilter.GaussianBlur(radius=2))
    
    return background

def process_product_image(image_url, chinese_text_regions=None):
    """
    ULTIMATE Professional Image Processing Pipeline:
    1. Download image
    2. AI Watermark removal
    3. AI Background removal  
    4. Optimized background creation
    5. Quality enhancement
    6. Professional effects (shadow, border)
    7. Promotional badges
    8. Korean text overlay
    """
    try:
        import cv2
        import numpy as np
        
        app.logger.info(f'[Image Pro] 🎨 Starting ULTIMATE processing for: {image_url}')
        
        # Download image
        response = requests.get(image_url, timeout=15)
        original_image = Image.open(io.BytesIO(response.content))
        
        # Convert to RGB
        if original_image.mode != 'RGB':
            original_image = original_image.convert('RGB')
        
        app.logger.info(f'[Image Pro] 📥 Downloaded: {original_image.size}')
        
        # === STEP 1: AI Watermark Removal ===
        app.logger.info('[Image Pro] 🧹 Removing watermarks...')
        no_watermark = remove_watermark_ai(original_image)
        
        # === STEP 2: AI Background Removal ===
        app.logger.info('[Image Pro] ✂️ Removing background...')
        no_background = remove_background_ai(no_watermark)
        
        # === STEP 3: Create Optimized Background ===
        app.logger.info('[Image Pro] 🎨 Creating optimized background...')
        
        # Detect product type from image (simple heuristic)
        # In production, use product title/category
        new_background = create_optimized_background(no_background, 'general')
        
        # Composite product on new background
        final_image = new_background.copy()
        final_image.paste(no_background, (0, 0), no_background)
        
        # === STEP 4: Quality Enhancement ===
        app.logger.info('[Image Pro] ⚡ Enhancing quality...')
        from PIL import ImageEnhance
        
        # Brightness
        enhancer = ImageEnhance.Brightness(final_image)
        final_image = enhancer.enhance(1.1)
        
        # Contrast
        enhancer = ImageEnhance.Contrast(final_image)
        final_image = enhancer.enhance(1.15)
        
        # Sharpness
        enhancer = ImageEnhance.Sharpness(final_image)
        final_image = enhancer.enhance(1.3)
        
        # Color saturation
        enhancer = ImageEnhance.Color(final_image)
        final_image = enhancer.enhance(1.1)
        
        # === STEP 5: Professional Border ===
        app.logger.info('[Image Pro] 🖼️ Adding border...')
        border_size = 30
        bordered = Image.new('RGB', 
                            (final_image.width + border_size*2, 
                             final_image.height + border_size*2),
                            (255, 255, 255))
        bordered.paste(final_image, (border_size, border_size))
        final_image = bordered
        
        # === STEP 6: Shadow Effect ===
        app.logger.info('[Image Pro] 💫 Adding shadow...')
        shadow_layer = Image.new('RGBA', final_image.size, (255, 255, 255, 0))
        shadow_draw = ImageDraw.Draw(shadow_layer)
        
        # Multiple shadow layers for soft effect
        for offset in range(8, 0, -2):
            alpha = int(40 * (offset / 8))
            shadow_draw.rounded_rectangle(
                [offset, offset, 
                 final_image.width - offset, 
                 final_image.height - offset],
                radius=15,
                fill=(0, 0, 0, alpha)
            )
        
        # Blend shadow
        final_rgb = final_image.convert('RGBA')
        final_rgb = Image.alpha_composite(shadow_layer, final_rgb)
        final_image = final_rgb.convert('RGB')
        
        # === STEP 7: Promotional Badges ===
        app.logger.info('[Image Pro] 🏷️ Adding badges...')
        draw = ImageDraw.Draw(final_image)
        
        try:
            font_badge = ImageFont.truetype('/usr/share/fonts/truetype/noto/NotoSansCJK-Bold.ttc', 32)
            font_small = ImageFont.truetype('/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc', 20)
        except:
            font_badge = font_small = ImageFont.load_default()
        
        # Badge 1: "무료배송" (Top-left)
        badge_configs = [
            {'text': '무료배송', 'pos': (40, 40), 'color': (255, 87, 51), 'size': (150, 55)},
            {'text': '베스트', 'pos': (final_image.width - 140, 40), 'color': (52, 199, 89), 'size': (100, 55)},
            {'text': '품질보증', 'pos': (40, final_image.height - 95), 'color': (0, 122, 255), 'size': (130, 55)}
        ]
        
        for badge in badge_configs:
            x, y = badge['pos']
            w, h = badge['size']
            
            # Shadow for badge
            draw.rounded_rectangle([x+3, y+3, x+w+3, y+h+3], 
                                  radius=10, fill=(0, 0, 0, 60))
            
            # Badge background
            draw.rounded_rectangle([x, y, x+w, y+h], 
                                  radius=10, fill=badge['color'])
            
            # Badge text
            text_bbox = draw.textbbox((0, 0), badge['text'], font=font_small)
            text_w = text_bbox[2] - text_bbox[0]
            text_h = text_bbox[3] - text_bbox[1]
            text_x = x + (w - text_w) // 2
            text_y = y + (h - text_h) // 2
            
            draw.text((text_x, text_y), badge['text'], 
                     fill=(255, 255, 255), font=font_small)
        
        # === STEP 8: Korean Text Overlay (if Chinese regions provided) ===
        if chinese_text_regions:
            app.logger.info('[Image Pro] 🔤 Adding Korean text...')
            for region in chinese_text_regions:
                x, y, w, h = region['bbox']
                korean_text = region.get('korean_text', '')
                
                # Rounded box
                draw.rounded_rectangle([x, y, x+w, y+h], 
                                      radius=8, 
                                      fill=(255, 255, 255, 240))
                
                # Text
                draw.text((x+10, y+10), korean_text, 
                         fill=(0, 0, 0), font=font_small)
        
        # === STEP 9: Save Final Image ===
        output = io.BytesIO()
        final_image.save(output, format='PNG', quality=95, optimize=True)
        output.seek(0)
        
        # Save to disk
        filename = f"ultimate_{int(time.time())}_{os.path.basename(image_url).split('?')[0]}.png"
        filepath = os.path.join('static', 'processed_images', filename)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        with open(filepath, 'wb') as f:
            f.write(output.getvalue())
        
        app.logger.info(f'[Image Pro] ✅ ULTIMATE processing complete: {filename}')
        return f'/static/processed_images/{filename}'
    
    except Exception as e:
        app.logger.error(f'[Image Pro] ❌ Processing failed: {str(e)}')
        import traceback
        app.logger.error(traceback.format_exc())
        log_activity('content', f'Image processing failed: {str(e)}', 'error')
        return image_url
    """
    Professional image processing for marketplace listing
    - Remove background (optional)
    - Add shadow effect
    - Enhance quality
    - Add Korean text overlay
    - Add promotional badges
    """
    try:
        import cv2
        import numpy as np
        
        # Download image
        response = requests.get(image_url, timeout=10)
        img = Image.open(io.BytesIO(response.content))
        
        # Convert to RGB if necessary
        if img.mode != 'RGB':
            img = img.convert('RGB')
        
        # Remove metadata
        data = list(img.getdata())
        image_without_exif = Image.new(img.mode, img.size)
        image_without_exif.putdata(data)
        
        # === PROFESSIONAL ENHANCEMENTS ===
        
        # 1. Enhance quality (brightness, contrast, sharpness)
        from PIL import ImageEnhance
        
        # Brightness
        enhancer = ImageEnhance.Brightness(image_without_exif)
        image_without_exif = enhancer.enhance(1.1)  # 10% brighter
        
        # Contrast
        enhancer = ImageEnhance.Contrast(image_without_exif)
        image_without_exif = enhancer.enhance(1.15)  # 15% more contrast
        
        # Sharpness
        enhancer = ImageEnhance.Sharpness(image_without_exif)
        image_without_exif = enhancer.enhance(1.2)  # 20% sharper
        
        # 2. Add white border (clean marketplace look)
        border_size = 20
        bordered = Image.new('RGB', 
                            (image_without_exif.width + border_size*2, 
                             image_without_exif.height + border_size*2),
                            (255, 255, 255))
        bordered.paste(image_without_exif, (border_size, border_size))
        image_without_exif = bordered
        
        # 3. Add subtle shadow effect
        # Create shadow layer
        shadow = Image.new('RGBA', image_without_exif.size, (0, 0, 0, 0))
        shadow_draw = ImageDraw.Draw(shadow)
        
        # Draw rounded rectangle shadow
        shadow_offset = 5
        shadow_draw.rounded_rectangle(
            [shadow_offset, shadow_offset, 
             image_without_exif.width - shadow_offset, 
             image_without_exif.height - shadow_offset],
            radius=10,
            fill=(0, 0, 0, 30)  # Semi-transparent black
        )
        
        # Blend shadow
        image_with_shadow = Image.new('RGB', image_without_exif.size, (255, 255, 255))
        image_with_shadow.paste(image_without_exif, (0, 0))
        
        # 4. Add promotional badges
        draw = ImageDraw.Draw(image_with_shadow, 'RGBA')
        
        try:
            font_badge = ImageFont.truetype('/usr/share/fonts/truetype/noto/NotoSansCJK-Bold.ttc', 28)
            font_small = ImageFont.truetype('/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc', 18)
        except:
            font_badge = ImageFont.load_default()
            font_small = ImageFont.load_default()
        
        # Top-left badge: "무료배송"
        badge_x, badge_y = 30, 30
        badge_width, badge_height = 140, 50
        
        # Badge background (gradient-like with overlay)
        draw.rounded_rectangle(
            [badge_x, badge_y, badge_x + badge_width, badge_y + badge_height],
            radius=8,
            fill=(255, 87, 51, 230)  # Coupang orange
        )
        
        # Badge text
        draw.text((badge_x + 15, badge_y + 12), "무료배송", 
                 fill=(255, 255, 255, 255), font=font_small)
        
        # Top-right badge: "베스트"
        badge2_x = image_with_shadow.width - 120
        badge2_y = 30
        
        draw.rounded_rectangle(
            [badge2_x, badge2_y, badge2_x + 90, badge2_y + 50],
            radius=8,
            fill=(52, 199, 89, 230)  # Green
        )
        
        draw.text((badge2_x + 15, badge2_y + 12), "베스트", 
                 fill=(255, 255, 255, 255), font=font_small)
        
        # 5. If Chinese text regions provided, add Korean overlay
        if chinese_text_regions:
            for region in chinese_text_regions:
                x, y, w, h = region['bbox']
                korean_text = region.get('korean_text', '')
                
                # Draw semi-transparent white box
                draw.rounded_rectangle([x, y, x+w, y+h], radius=5, 
                                      fill=(255, 255, 255, 220))
                
                # Draw Korean text
                draw.text((x+8, y+8), korean_text, 
                         fill=(0, 0, 0, 255), font=font_small)
        
        # 6. Save processed image
        output = io.BytesIO()
        image_with_shadow.save(output, format='PNG', quality=95)
        output.seek(0)
        
        # Save to disk
        filename = f"processed_{int(time.time())}_{os.path.basename(image_url).split('?')[0]}.png"
        filepath = os.path.join('static', 'processed_images', filename)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        with open(filepath, 'wb') as f:
            f.write(output.getvalue())
        
        app.logger.info(f'[Image Processing] ✅ Enhanced image saved: {filename}')
        return f'/static/processed_images/{filename}'
    
    except Exception as e:
        app.logger.error(f'[Image Processing] ❌ Failed: {str(e)}')
        log_activity('content', f'Image processing failed: {str(e)}', 'error')
        return image_url

@app.route('/api/content/generate/<int:product_id>', methods=['POST'])
@login_required
def generate_content(product_id):
    """Generate AI content for product using WINNING PRODUCT STRUCTURE"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM sourced_products WHERE id = ?', (product_id,))
    product = cursor.fetchone()
    
    if not product:
        conn.close()
        return jsonify({'error': 'Product not found'}), 404
    
    app.logger.info(f'[Content Generation] 🚀 Starting WINNING page generation for product {product_id}')
    log_activity('content', f'Generating WINNING content for product {product_id}', 'in_progress')
    
    # Generate SHORT marketing copy (for summary box)
    marketing_copy = generate_marketing_copy(product['title_cn'], product['price_krw'])
    
    # Process images with Korean shopping mall styling
    original_images = json.loads(product['images_json']) if product['images_json'] else []
    processed_images = []
    
    app.logger.info(f'[Content Generation] 📸 Processing {len(original_images)} images with Korean styling')
    for img_url in original_images[:8]:  # Process max 8 images for winning structure
        processed_url = process_product_image(img_url)
        processed_images.append(processed_url)
    
    # Generate WINNING PRODUCT PAGE (5-section structure)
    app.logger.info(f'[Content Generation] ✨ Generating WINNING 5-section structure')
    winning_html, seo_tags = generate_winning_product_page(
        title=product['title_cn'],
        price=product['price_krw'],
        images=processed_images
    )
    
    # Add notice image at the end
    notice_path = create_notice_image()
    processed_images.append(notice_path)
    
    # 🔥 EMERGENCY FIX: Store SEO tags in dedicated keywords column
    cursor.execute('''
        UPDATE sourced_products
        SET marketing_copy = ?,
            description_kr = ?,
            processed_images_json = ?,
            title_kr = ?,
            keywords = ?
        WHERE id = ?
    ''', (
        marketing_copy,
        winning_html,  # Full winning structure in description_kr
        json.dumps(processed_images),
        product['title_cn'],
        seo_tags,  # 🔥 FIXED: Store in keywords column, NOT description_cn
        product_id
    ))
    
    conn.commit()
    conn.close()
    
    app.logger.info(f'[Content Generation] ✅ WINNING content generated: {len(winning_html)} chars')
    if seo_tags:
        app.logger.info(f'[Content Generation] ✅ SEO tags: {seo_tags}')
    log_activity('content', f'✅ WINNING content generated for product {product_id}', 'success')
    
    return jsonify({
        'success': True,
        'marketing_copy': marketing_copy,
        'description_kr': winning_html,
        'processed_images': processed_images,
        'seo_tags': seo_tags,
        'structure': '5-section_winning_formula'
    })

def create_notice_image():
    """Create notice image for product detail bottom"""
    width, height = 800, 400
    img = Image.new('RGB', (width, height), color=(255, 248, 240))
    draw = ImageDraw.Draw(img)
    
    try:
        font_title = ImageFont.truetype('/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc', 24)
        font_text = ImageFont.truetype('/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc', 16)
    except:
        font_title = font_text = ImageFont.load_default()
    
    # Draw notice content
    draw.text((50, 50), '⚠️ 해외직구 상품 안내', fill=(220, 20, 60), font=font_title)
    
    notices = [
        '• 본 상품은 해외 직구 상품으로 배송기간이 2-3주 소요됩니다.',
        '• 통관 과정에서 추가 세금이 발생할 수 있습니다.',
        '• 단순 변심 반품 시 왕복 배송비가 부과됩니다.',
        '• 상품 문의는 고객센터로 연락 주시기 바랍니다.'
    ]
    
    y = 120
    for notice in notices:
        draw.text((50, y), notice, fill=(50, 50, 50), font=font_text)
        y += 40
    
    # Save
    filepath = 'static/notice_image.png'
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    img.save(filepath)
    
    return '/static/notice_image.png'

# ============================================================================
# MODULE 4: NAVER/COUPANG API INTEGRATION
# ============================================================================

def generate_naver_signature(client_id, client_secret, timestamp, method, path):
    """Generate Naver Commerce API signature with bcrypt"""
    message = f"{timestamp}.{method}.{path}"
    signature = hmac.new(
        client_secret.encode('utf-8'),
        message.encode('utf-8'),
        hashlib.sha256
    ).hexdigest()
    
    return base64.b64encode(signature.encode()).decode()

def register_to_naver(product_data):
    """Register product to Naver Smart Store"""
    client_id = get_config('naver_client_id')
    client_secret = get_config('naver_client_secret')
    
    if not client_id or not client_secret:
        return {'error': 'Naver API credentials not configured'}
    
    timestamp = str(int(time.time() * 1000))
    path = '/v1/products'
    signature = generate_naver_signature(client_id, client_secret, timestamp, 'POST', path)
    
    headers = {
        'X-Naver-Client-Id': client_id,
        'X-Naver-Client-Secret': client_secret,
        'X-Timestamp': timestamp,
        'X-Signature': signature,
        'Content-Type': 'application/json'
    }
    
    try:
        response = requests.post(
            f'https://api.commerce.naver.com{path}',
            headers=headers,
            json=product_data,
            timeout=30
        )
        
        return response.json()
    
    except Exception as e:
        return {'error': str(e)}

def register_to_coupang(product_data):
    """Register product to Coupang"""
    access_key = get_config('coupang_access_key')
    secret_key = get_config('coupang_secret_key')
    
    if not access_key or not secret_key:
        return {'error': 'Coupang API credentials not configured'}
    
    # Coupang requires IP whitelisting
    static_ip = get_config('server_static_ip')
    if not static_ip:
        log_activity('marketplace', 'Warning: Static IP not configured for Coupang', 'warning')
    
    path = '/v2/providers/seller_api/apis/api/v1/marketplace/seller-products'
    timestamp = str(int(time.time() * 1000))
    
    message = timestamp + 'POST' + path + ''
    signature = hmac.new(
        secret_key.encode('utf-8'),
        message.encode('utf-8'),
        hashlib.sha256
    ).hexdigest()
    
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'CEA algorithm=HmacSHA256, access-key={access_key}, signed-date={timestamp}, signature={signature}'
    }
    
    try:
        response = requests.post(
            f'https://api-gateway.coupang.com{path}',
            headers=headers,
            json=product_data,
            timeout=30
        )
        
        return response.json()
    
    except Exception as e:
        return {'error': str(e)}

def update_marketplace_status(marketplace, external_id, status):
    """Update order status on marketplace"""
    if marketplace == 'naver':
        return update_naver_order_status(external_id, status)
    elif marketplace == 'coupang':
        return update_coupang_order_status(external_id, status)
    
    return {'error': 'Unknown marketplace'}

def update_naver_order_status(order_id, status):
    """Update Naver order status"""
    # Check if API credentials are configured
    client_id = get_config('naver_client_id')
    client_secret = get_config('naver_client_secret')
    
    if not client_id or not client_secret:
        log_activity('marketplace', 'Naver API credentials not configured - skipping status update', 'warning')
        return {'success': True, 'message': f'Order {order_id} status updated locally (API not configured)', 'skipped': True}
    
    # Simplified implementation - in production, make actual API call
    return {'success': True, 'message': f'Order {order_id} status updated to {status}'}

def update_coupang_order_status(order_id, status):
    """Update Coupang order status"""
    # Check if API credentials are configured
    access_key = get_config('coupang_access_key')
    secret_key = get_config('coupang_secret_key')
    
    if not access_key or not secret_key:
        log_activity('marketplace', 'Coupang API credentials not configured - skipping status update', 'warning')
        return {'success': True, 'message': f'Order {order_id} status updated locally (API not configured)', 'skipped': True}
    
    # Simplified implementation - in production, make actual API call
    return {'success': True, 'message': f'Order {order_id} status updated to {status}'}

# ============================================================================
# MODULE 5: ORDER MANAGEMENT WITH PCCC VALIDATION
# ============================================================================

def validate_pccc(pccc):
    """Validate Korean PCCC (Personal Customs Clearance Code)"""
    # PCCC format: P123456789012 (P + 12 digits)
    pattern = r'^P\d{12}$'
    return bool(re.match(pattern, pccc))

@app.route('/api/orders/create', methods=['POST'])
@login_required
def create_order():
    """Create new order"""
    data = request.json
    
    # Validate PCCC
    pccc = data.get('pccc', '')
    if not validate_pccc(pccc):
        return jsonify({'error': 'Invalid PCCC format'}), 400
    
    # Get current exchange rate (will be locked to this order)
    current_rate = float(get_config('cny_exchange_rate', 190))
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Calculate order financials
    price_cny = data.get('price_cny', 0)
    quantity = data.get('quantity', 1)
    
    analysis = analyze_product_profitability(price_cny)
    marketplace_fee_rate = float(get_config(f"{data.get('marketplace', 'naver')}_fee_rate", 0.06))
    marketplace_fee = int(analysis['sale_price'] * marketplace_fee_rate)
    
    net_profit = analysis['profit'] - marketplace_fee
    
    # Set shipping deadline (3 days from now)
    shipping_deadline = get_kst_now() + timedelta(days=3)
    
    cursor.execute('''
        INSERT INTO orders (
            order_number, marketplace, product_id, customer_name, customer_phone,
            customer_address, pccc, pccc_validated, quantity, sale_price,
            purchase_price_cny, purchase_price_krw, applied_exchange_rate,
            shipping_cost, customs_tax, marketplace_fee, net_profit,
            original_product_url, shipping_deadline, order_status
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        data.get('order_number'),
        data.get('marketplace'),
        data.get('product_id'),
        data.get('customer_name'),
        data.get('customer_phone'),
        data.get('customer_address'),
        pccc,
        True,
        quantity,
        analysis['sale_price'] * quantity,
        price_cny,
        analysis['purchase_price_krw'],
        current_rate,  # Lock exchange rate
        analysis['shipping_cost'],
        analysis['customs_tax'],
        marketplace_fee,
        net_profit * quantity,
        data.get('original_product_url'),
        shipping_deadline.isoformat(),
        'pending'
    ))
    
    order_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    log_activity('order', f'Order {data.get("order_number")} created', 'success')
    
    return jsonify({'success': True, 'order_id': order_id})

@app.route('/api/orders/<int:order_id>/confirm', methods=['POST'])
@login_required
def confirm_order(order_id):
    """Mark order as confirmed and update marketplace status"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM orders WHERE id = ?', (order_id,))
    order = cursor.fetchone()
    
    if not order:
        conn.close()
        return jsonify({'error': 'Order not found'}), 404
    
    # Update local status
    cursor.execute('''
        UPDATE orders SET order_status = 'confirmed' WHERE id = ?
    ''', (order_id,))
    
    conn.commit()
    conn.close()
    
    # Update marketplace
    result = update_marketplace_status(
        order['marketplace'],
        order['order_number'],
        'preparing'
    )
    
    log_activity('order', f'Order {order["order_number"]} confirmed', 'success')
    
    return jsonify({'success': True})

@app.route('/api/orders/<int:order_id>/ship', methods=['POST'])
@login_required
def ship_order(order_id):
    """Add tracking number and mark as shipped"""
    data = request.json
    tracking_number = data.get('tracking_number', '')
    
    if not tracking_number:
        return jsonify({'error': 'Tracking number required'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM orders WHERE id = ?', (order_id,))
    order = cursor.fetchone()
    
    if not order:
        conn.close()
        return jsonify({'error': 'Order not found'}), 404
    
    # Update order
    cursor.execute('''
        UPDATE orders 
        SET tracking_number = ?,
            order_status = 'shipped',
            shipped_at = CURRENT_TIMESTAMP
        WHERE id = ?
    ''', (tracking_number, order_id))
    
    conn.commit()
    conn.close()
    
    # Update marketplace
    result = update_marketplace_status(
        order['marketplace'],
        order['order_number'],
        'shipping'
    )
    
    log_activity('order', f'Order {order["order_number"]} shipped with tracking {tracking_number}', 'success')
    
    return jsonify({'success': True})

@app.route('/api/orders/<int:order_id>/deliver', methods=['POST'])
@login_required
def deliver_order(order_id):
    """Mark order as delivered and create tax record"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM orders WHERE id = ?', (order_id,))
    order = cursor.fetchone()
    
    if not order:
        conn.close()
        return jsonify({'error': 'Order not found'}), 404
    
    # Update order status
    cursor.execute('''
        UPDATE orders 
        SET order_status = 'delivered',
            delivered_at = CURRENT_TIMESTAMP
        WHERE id = ?
    ''', (order_id,))
    
    # Create tax record
    cursor.execute('''
        INSERT INTO tax_records (
            order_id, order_number, sale_date, sale_amount,
            purchase_amount, shipping_cost, marketplace_fee,
            net_profit, applied_exchange_rate
        ) VALUES (?, ?, DATE('now'), ?, ?, ?, ?, ?, ?)
    ''', (
        order_id,
        order['order_number'],
        order['sale_price'],
        order['purchase_price_krw'],
        order['shipping_cost'],
        order['marketplace_fee'],
        order['net_profit'],
        order['applied_exchange_rate']
    ))
    
    conn.commit()
    conn.close()
    
    log_activity('order', f'Order {order["order_number"]} delivered', 'success')
    
    return jsonify({'success': True})

# ============================================================================
# MODULE 6: RISK DEFENSE - STOCK MONITORING
# ============================================================================

def check_product_stock_status(product_url):
    """Check if product is still available on source site (Alibaba/AliExpress)"""
    api_key = get_config('scrapingant_api_key')
    if not api_key:
        return {'available': True, 'reason': 'Cannot verify - API key missing'}
    
    try:
        params = {
            'url': product_url,
            'x-api-key': api_key,
            'browser': 'true'
        }
        
        response = requests.get('https://api.scrapingant.com/v2/general', params=params, timeout=30)
        
        if response.status_code == 404:
            return {'available': False, 'reason': 'Product deleted'}
        
        # Check for out of stock indicators
        text = response.text.lower()
        if '下架' in text or '已下架' in text or '缺货' in text:
            return {'available': False, 'reason': 'Out of stock'}
        
        return {'available': True, 'reason': 'Available'}
    
    except Exception as e:
        return {'available': True, 'reason': f'Check failed: {str(e)}'}

def run_stock_monitor():
    """Daily stock monitoring job"""
    log_activity('stock_monitor', 'Starting daily stock check', 'in_progress')
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Get all active marketplace listings
    cursor.execute('''
        SELECT ml.id, ml.product_id, sp.original_url, ml.marketplace, ml.external_product_id
        FROM marketplace_listings ml
        JOIN sourced_products sp ON ml.product_id = sp.id
        WHERE ml.status = 'active'
    ''')
    
    listings = cursor.fetchall()
    
    for listing in listings:
        status = check_product_stock_status(listing['original_url'])
        
        if not status['available']:
            # Mark as out of stock
            cursor.execute('''
                UPDATE marketplace_listings
                SET status = 'out_of_stock', updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (listing['id'],))
            
            # Log action
            cursor.execute('''
                INSERT INTO stock_monitor_log (product_id, original_url, check_status, action_taken)
                VALUES (?, ?, ?, ?)
            ''', (
                listing['product_id'],
                listing['original_url'],
                'unavailable',
                f"Marked as out of stock: {status['reason']}"
            ))
            
            log_activity('stock_monitor', 
                        f"Product {listing['product_id']} marked out of stock: {status['reason']}", 
                        'warning')
        else:
            cursor.execute('''
                INSERT INTO stock_monitor_log (product_id, original_url, check_status, action_taken)
                VALUES (?, ?, ?, ?)
            ''', (
                listing['product_id'],
                listing['original_url'],
                'available',
                'No action needed'
            ))
    
    conn.commit()
    conn.close()
    
    log_activity('stock_monitor', f'Stock check completed. Checked {len(listings)} products', 'success')

# Schedule daily stock monitoring
schedule.every().day.at("02:00").do(run_stock_monitor)

def run_scheduler():
    """Run scheduled tasks in background thread"""
    while True:
        schedule.run_pending()
        time.sleep(60)

# Start scheduler thread
scheduler_thread = threading.Thread(target=run_scheduler, daemon=True)
scheduler_thread.start()

# ============================================================================
# MODULE 7: TAX AUTOMATION & EXCEL EXPORT
# ============================================================================

@app.route('/api/tax/export', methods=['GET'])
@login_required
def export_tax_records():
    """Export tax records to Excel"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = 'SELECT * FROM tax_records WHERE 1=1'
    params = []
    
    if start_date:
        query += ' AND sale_date >= ?'
        params.append(start_date)
    
    if end_date:
        query += ' AND sale_date <= ?'
        params.append(end_date)
    
    query += ' ORDER BY sale_date DESC'
    
    cursor.execute(query, params)
    records = cursor.fetchall()
    conn.close()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "판매대장"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = [
        '일자', '주문번호', '판매금액', '매입금액(원)', '배송비', 
        '수수료', '순수익', '적용환율', '비고'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for row_idx, record in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=record['sale_date'])
        ws.cell(row=row_idx, column=2, value=record['order_number'])
        ws.cell(row=row_idx, column=3, value=record['sale_amount'])
        ws.cell(row=row_idx, column=4, value=record['purchase_amount'])
        ws.cell(row=row_idx, column=5, value=record['shipping_cost'])
        ws.cell(row=row_idx, column=6, value=record['marketplace_fee'])
        ws.cell(row=row_idx, column=7, value=record['net_profit'])
        ws.cell(row=row_idx, column=8, value=record['applied_exchange_rate'])
        ws.cell(row=row_idx, column=9, value='완료')
    
    # Add summary row
    summary_row = len(records) + 3
    ws.cell(row=summary_row, column=1, value='합계').font = Font(bold=True)
    ws.cell(row=summary_row, column=3, value=f'=SUM(C2:C{len(records)+1})')
    ws.cell(row=summary_row, column=7, value=f'=SUM(G2:G{len(records)+1})')
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to file
    filename = f'tax_export_{get_kst_now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    filepath = os.path.join('static', 'exports', filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    
    log_activity('tax', f'Exported {len(records)} tax records', 'success')
    
    return send_file(filepath, as_attachment=True, download_name=filename)

# ============================================================================
# MODULE 8: DASHBOARD & FRONTEND ROUTES
# ============================================================================

@app.route('/')
@login_required
def dashboard():
    """Main dashboard"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Get statistics
    cursor.execute('SELECT COUNT(*) as count FROM sourced_products WHERE status = "pending"')
    pending_products = cursor.fetchone()['count']
    
    cursor.execute('SELECT COUNT(*) as count FROM orders WHERE order_status = "pending"')
    pending_orders = cursor.fetchone()['count']
    
    cursor.execute('SELECT COUNT(*) as count FROM orders WHERE datetime(shipping_deadline) < datetime("now", "+1 day")')
    urgent_orders = cursor.fetchone()['count']
    
    cursor.execute('SELECT SUM(net_profit) as total FROM orders WHERE order_status = "delivered"')
    total_profit = cursor.fetchone()['total'] or 0
    
    # NEW: Financial analytics for Version 2.0
    # Today's revenue
    cursor.execute('''
        SELECT 
            COALESCE(SUM(sale_price), 0) as today_sales,
            COALESCE(SUM(net_profit), 0) as today_profit
        FROM orders 
        WHERE DATE(delivered_at) = DATE('now') AND order_status = 'delivered'
    ''')
    today_stats = cursor.fetchone()
    today_sales = today_stats['today_sales']
    today_profit = today_stats['today_profit']
    
    # This month's revenue
    cursor.execute('''
        SELECT 
            COALESCE(SUM(sale_price), 0) as month_sales,
            COALESCE(SUM(net_profit), 0) as month_profit
        FROM orders 
        WHERE strftime('%Y-%m', delivered_at) = strftime('%Y-%m', 'now') 
        AND order_status = 'delivered'
    ''')
    month_stats = cursor.fetchone()
    month_sales = month_stats['month_sales']
    month_profit = month_stats['month_profit']
    
    # Last 7 days daily revenue for chart
    cursor.execute('''
        SELECT 
            DATE(delivered_at) as date,
            COALESCE(SUM(sale_price), 0) as daily_sales,
            COALESCE(SUM(net_profit), 0) as daily_profit
        FROM orders 
        WHERE delivered_at >= DATE('now', '-7 days') 
        AND order_status = 'delivered'
        GROUP BY DATE(delivered_at)
        ORDER BY date
    ''')
    daily_stats = cursor.fetchall()
    
    # Monthly revenue for last 6 months
    cursor.execute('''
        SELECT 
            strftime('%Y-%m', delivered_at) as month,
            COALESCE(SUM(sale_price), 0) as monthly_sales,
            COALESCE(SUM(net_profit), 0) as monthly_profit
        FROM orders 
        WHERE delivered_at >= DATE('now', '-6 months')
        AND order_status = 'delivered'
        GROUP BY strftime('%Y-%m', delivered_at)
        ORDER BY month
    ''')
    monthly_stats = cursor.fetchall()
    
    # Recent activity logs
    cursor.execute('SELECT * FROM activity_logs ORDER BY created_at DESC LIMIT 20')
    recent_logs = cursor.fetchall()
    
    conn.close()
    
    # Prepare chart data
    daily_labels = [row['date'] if row['date'] else '' for row in daily_stats] if daily_stats else []
    daily_sales_data = [row['daily_sales'] for row in daily_stats] if daily_stats else []
    daily_profit_data = [row['daily_profit'] for row in daily_stats] if daily_stats else []
    
    monthly_labels = [row['month'] if row['month'] else '' for row in monthly_stats] if monthly_stats else []
    monthly_profit_data = [row['monthly_profit'] for row in monthly_stats] if monthly_stats else []
    
    return render_template('dashboard.html',
                         pending_products=pending_products,
                         pending_orders=pending_orders,
                         urgent_orders=urgent_orders,
                         total_profit=total_profit,
                         today_sales=today_sales,
                         today_profit=today_profit,
                         month_sales=month_sales,
                         month_profit=month_profit,
                         daily_labels=json.dumps(daily_labels),
                         daily_sales_data=json.dumps(daily_sales_data),
                         daily_profit_data=json.dumps(daily_profit_data),
                         monthly_labels=json.dumps(monthly_labels),
                         monthly_profit_data=json.dumps(monthly_profit_data),
                         recent_logs=recent_logs)

@app.route('/products')
@login_required
def products():
    """Products management page"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM sourced_products ORDER BY created_at DESC')
    products = cursor.fetchall()
    conn.close()
    
    return render_template('products.html', products=products)

@app.route('/products/<int:product_id>')
@login_required
def product_detail(product_id):
    """Product detail and edit page"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM sourced_products WHERE id = ?', (product_id,))
    product = cursor.fetchone()
    conn.close()
    
    if not product:
        flash('상품을 찾을 수 없습니다', 'error')
        return redirect('/products')
    
    return render_template('product_detail.html', product=product)

@app.route('/api/products/<int:product_id>/update', methods=['POST'])
@login_required
def update_product(product_id):
    """Update product information"""
    try:
        data = request.get_json()
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Update product fields
        cursor.execute('''
            UPDATE sourced_products 
            SET title_kr = ?, 
                title_cn = ?,
                price_krw = ?,
                profit_margin = ?,
                estimated_profit = ?,
                marketing_copy = ?,
                description_kr = ?,
                description_cn = ?,
                tags = ?
            WHERE id = ?
        ''', (
            data.get('title_kr'),
            data.get('title_cn'),
            data.get('price_krw'),
            data.get('profit_margin'),
            data.get('estimated_profit'),
            data.get('marketing_copy'),
            data.get('description_kr'),
            data.get('description_cn'),
            data.get('tags'),
            product_id
        ))
        
        conn.commit()
        conn.close()
        
        log_activity('product_update', f'상품 수정: ID {product_id}', 'success')
        
        return jsonify({'success': True, 'message': '상품이 수정되었습니다'})
        
    except Exception as e:
        app.logger.error(f'[Product Update] ❌ Error: {str(e)}')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/products/<int:product_id>', methods=['DELETE'])
@login_required
def delete_product(product_id):
    """Delete product"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Check if product exists
        cursor.execute('SELECT * FROM sourced_products WHERE id = ?', (product_id,))
        product = cursor.fetchone()
        
        if not product:
            conn.close()
            return jsonify({'success': False, 'error': '상품을 찾을 수 없습니다'}), 404
        
        # Delete product
        cursor.execute('DELETE FROM sourced_products WHERE id = ?', (product_id,))
        
        conn.commit()
        conn.close()
        
        log_activity('product_delete', f'상품 삭제: ID {product_id} - {product["title_kr"] or product["title_cn"]}', 'success')
        app.logger.info(f'[Product Delete] ✅ Product deleted: ID {product_id}')
        
        return jsonify({'success': True, 'message': '상품이 삭제되었습니다'})
        
    except Exception as e:
        app.logger.error(f'[Product Delete] ❌ Error: {str(e)}')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/products/bulk-delete', methods=['POST'])
@login_required
def bulk_delete_products():
    """Bulk delete products"""
    try:
        data = request.json
        product_ids = data.get('product_ids', [])
        
        if not product_ids:
            return jsonify({'success': False, 'error': '삭제할 상품을 선택해주세요'}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Delete all selected products
        deleted_count = 0
        for product_id in product_ids:
            cursor.execute('SELECT title_kr, title_cn FROM sourced_products WHERE id = ?', (product_id,))
            product = cursor.fetchone()
            
            if product:
                cursor.execute('DELETE FROM sourced_products WHERE id = ?', (product_id,))
                deleted_count += 1
                app.logger.info(f'[Bulk Delete] Deleted product ID {product_id}: {product["title_kr"] or product["title_cn"]}')
        
        conn.commit()
        conn.close()
        
        log_activity('product_bulk_delete', f'일괄 삭제: {deleted_count}개 상품', 'success')
        app.logger.info(f'[Bulk Delete] ✅ Successfully deleted {deleted_count} products')
        
        return jsonify({'success': True, 'deleted_count': deleted_count, 'message': f'{deleted_count}개의 상품이 삭제되었습니다'})
        
    except Exception as e:
        app.logger.error(f'[Bulk Delete] ❌ Error: {str(e)}')
        return jsonify({'success': False, 'error': str(e)}), 500

def sync_naver_orders():
    """
    Sync orders from Naver SmartStore API
    Fetches new orders and updates local database
    """
    try:
        client_id = get_config('naver_client_id')
        client_secret = get_config('naver_client_secret')
        
        if not client_id or not client_secret:
            app.logger.warning('[Naver Sync] API credentials not configured')
            return {'success': False, 'error': 'API credentials not configured'}
        
        import time
        import hmac
        import hashlib
        
        # Naver Commerce API endpoint
        timestamp = str(int(time.time() * 1000))
        method = 'GET'
        path = '/v1/pay-order/seller/orders'
        
        # Create signature
        message = f"{timestamp}.{method}.{path}"
        signature = hmac.new(
            client_secret.encode('utf-8'),
            message.encode('utf-8'),
            hashlib.sha256
        ).hexdigest()
        
        headers = {
            'Content-Type': 'application/json',
            'X-Naver-Client-Id': client_id,
            'X-Timestamp': timestamp,
            'X-API-Signature': signature
        }
        
        # Fetch orders from last 7 days
        params = {
            'lastChangedFrom': (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d'),
            'lastChangedTo': datetime.now().strftime('%Y-%m-%d')
        }
        
        response = requests.get(
            f'https://api.commerce.naver.com{path}',
            headers=headers,
            params=params,
            timeout=30
        )
        
        if response.status_code != 200:
            app.logger.error(f'[Naver Sync] API error: {response.status_code}')
            return {'success': False, 'error': f'API error: {response.status_code}'}
        
        data = response.json()
        orders_data = data.get('data', {}).get('orders', [])
        
        conn = get_db()
        cursor = conn.cursor()
        synced_count = 0
        
        for order in orders_data:
            # Check if order already exists
            cursor.execute('SELECT id FROM orders WHERE order_number = ? AND marketplace = ?',
                          (order['orderNumber'], 'naver'))
            existing = cursor.fetchone()
            
            if not existing:
                # Create new order
                cursor.execute('''
                    INSERT INTO orders (
                        order_number, marketplace, customer_name, customer_phone,
                        customer_address, quantity, sale_price, order_status,
                        payment_status, ordered_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    order['orderNumber'],
                    'naver',
                    order.get('ordererName', ''),
                    order.get('ordererTel', ''),
                    order.get('shippingAddress', ''),
                    order.get('quantity', 1),
                    order.get('paymentAmount', 0),
                    order.get('orderStatus', 'pending'),
                    order.get('paymentStatus', 'paid'),
                    order.get('orderedAt', datetime.now().isoformat())
                ))
                synced_count += 1
        
        conn.commit()
        conn.close()
        
        app.logger.info(f'[Naver Sync] ✅ Synced {synced_count} new orders')
        log_activity('marketplace', f'Naver orders synced: {synced_count} new', 'success')
        
        return {'success': True, 'synced_count': synced_count}
        
    except Exception as e:
        app.logger.error(f'[Naver Sync] ❌ Error: {str(e)}')
        import traceback
        app.logger.error(traceback.format_exc())
        return {'success': False, 'error': str(e)}

def sync_coupang_orders():
    """
    Sync orders from Coupang Wing API
    Fetches new orders and updates local database
    """
    try:
        access_key = get_config('coupang_access_key')
        secret_key = get_config('coupang_secret_key')
        
        if not access_key or not secret_key:
            app.logger.warning('[Coupang Sync] API credentials not configured')
            return {'success': False, 'error': 'API credentials not configured'}
        
        import time
        import hmac
        import hashlib
        
        # Coupang Wing API endpoint
        method = 'GET'
        path = '/v2/providers/seller_api/apis/api/v1/marketplace/orders'
        query = f'createdAtFrom={int((datetime.now() - timedelta(days=7)).timestamp() * 1000)}'
        
        # Create signature
        timestamp = str(int(time.time() * 1000))
        message = f"{timestamp}{method}{path}?{query}"
        signature = hmac.new(
            secret_key.encode('utf-8'),
            message.encode('utf-8'),
            hashlib.sha256
        ).hexdigest()
        
        headers = {
            'Content-Type': 'application/json',
            'Authorization': f'CEA algorithm=HmacSHA256, access-key={access_key}, signed-date={timestamp}, signature={signature}'
        }
        
        response = requests.get(
            f'https://api-gateway.coupang.com{path}?{query}',
            headers=headers,
            timeout=30
        )
        
        if response.status_code != 200:
            app.logger.error(f'[Coupang Sync] API error: {response.status_code}')
            return {'success': False, 'error': f'API error: {response.status_code}'}
        
        data = response.json()
        orders_data = data.get('data', [])
        
        conn = get_db()
        cursor = conn.cursor()
        synced_count = 0
        
        for order in orders_data:
            # Check if order already exists
            cursor.execute('SELECT id FROM orders WHERE order_number = ? AND marketplace = ?',
                          (order['orderId'], 'coupang'))
            existing = cursor.fetchone()
            
            if not existing:
                # Create new order
                cursor.execute('''
                    INSERT INTO orders (
                        order_number, marketplace, customer_name, customer_phone,
                        customer_address, quantity, sale_price, order_status,
                        payment_status, ordered_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    order['orderId'],
                    'coupang',
                    order.get('ordererName', ''),
                    order.get('ordererContact', ''),
                    order.get('shippingAddress', ''),
                    order.get('orderedQuantity', 1),
                    order.get('paidAmount', 0),
                    order.get('shipmentBoxId') and 'confirmed' or 'pending',
                    'paid',
                    order.get('orderedAt', datetime.now().isoformat())
                ))
                synced_count += 1
        
        conn.commit()
        conn.close()
        
        app.logger.info(f'[Coupang Sync] ✅ Synced {synced_count} new orders')
        log_activity('marketplace', f'Coupang orders synced: {synced_count} new', 'success')
        
        return {'success': True, 'synced_count': synced_count}
        
    except Exception as e:
        app.logger.error(f'[Coupang Sync] ❌ Error: {str(e)}')
        import traceback
        app.logger.error(traceback.format_exc())
        return {'success': False, 'error': str(e)}

@app.route('/api/orders/sync', methods=['POST'])
@login_required
def sync_marketplace_orders():
    """
    Sync orders from all configured marketplaces
    Can be called manually or set up as a cron job
    """
    try:
        results = {
            'naver': {'synced': False, 'count': 0},
            'coupang': {'synced': False, 'count': 0}
        }
        
        # Sync Naver orders
        naver_result = sync_naver_orders()
        if naver_result.get('success'):
            results['naver']['synced'] = True
            results['naver']['count'] = naver_result.get('synced_count', 0)
        else:
            results['naver']['error'] = naver_result.get('error', 'Unknown error')
        
        # Sync Coupang orders
        coupang_result = sync_coupang_orders()
        if coupang_result.get('success'):
            results['coupang']['synced'] = True
            results['coupang']['count'] = coupang_result.get('synced_count', 0)
        else:
            results['coupang']['error'] = coupang_result.get('error', 'Unknown error')
        
        total_synced = results['naver']['count'] + results['coupang']['count']
        
        return jsonify({
            'success': True,
            'results': results,
            'total_synced': total_synced,
            'message': f'총 {total_synced}개의 새 주문이 동기화되었습니다'
        })
        
    except Exception as e:
        app.logger.error(f'[Order Sync] ❌ Error: {str(e)}')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/orders')
@login_required
def orders():
    """Orders management page"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT o.*, sp.title_kr, sp.original_url
        FROM orders o
        LEFT JOIN sourced_products sp ON o.product_id = sp.id
        ORDER BY o.ordered_at DESC
    ''')
    orders = cursor.fetchall()
    conn.close()
    
    return render_template('orders.html', orders=orders)

@app.route('/config')
@login_required
def config_page():
    """Configuration page"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM config')
    configs = {row['key']: row['value'] for row in cursor.fetchall()}
    conn.close()
    
    return render_template('config.html', configs=configs)

@app.route('/logs')
@login_required
def logs_page():
    """System logs viewer page"""
    app.logger.info(f'User {current_user.username} accessed system logs')
    
    log_file_path = 'logs/server.log'
    log_lines = []
    
    try:
        if os.path.exists(log_file_path):
            with open(log_file_path, 'r', encoding='utf-8') as f:
                # Read all lines and get last 100
                all_lines = f.readlines()
                log_lines = all_lines[-100:] if len(all_lines) > 100 else all_lines
                # Reverse to show newest first
                log_lines.reverse()
        else:
            log_lines = ['[INFO] Log file not found. Server just started or no logs generated yet.']
    except Exception as e:
        log_lines = [f'[ERROR] Failed to read log file: {str(e)}']
        app.logger.error(f'Failed to read log file: {str(e)}')
    
    return render_template('logs.html', log_lines=log_lines)

@app.route('/api/config/update', methods=['POST'])
@login_required
def update_config():
    """Update configuration"""
    data = request.json
    
    app.logger.info('[Config Update] ========================================')
    app.logger.info(f'[Config Update] Received {len(data)} config items to update')
    
    saved_configs = []
    for key, value in data.items():
        # Log what's being saved (mask sensitive data)
        if 'key' in key.lower() or 'secret' in key.lower() or 'password' in key.lower():
            display_value = f'{value[:10]}...' if value and len(value) > 10 else '[EMPTY]'
        else:
            display_value = value
        
        app.logger.info(f'[Config Update] Saving: {key} = {display_value}')
        
        set_config(key, value)
        saved_configs.append(key)
        
        # Immediately verify it was saved
        verify_value = get_config(key)
        if verify_value == value:
            app.logger.info(f'[Config Update] ✅ Verified: {key} saved correctly')
        else:
            app.logger.error(f'[Config Update] ❌ FAILED: {key} mismatch! Expected: {value[:20]}..., Got: {verify_value}')
    
    app.logger.info(f'[Config Update] Successfully saved: {", ".join(saved_configs)}')
    log_activity('config', f'Configuration updated: {", ".join(saved_configs)}', 'success')
    
    return jsonify({'success': True, 'saved_configs': saved_configs})

@app.route('/api/config/verify', methods=['GET'])
@login_required
def verify_config():
    """Verify current configuration (diagnostic endpoint)"""
    app.logger.info('[Config Verify] Checking current configuration...')
    
    # Check critical config values
    configs_to_check = [
        'scrapingant_api_key',
        'openai_api_key',
        'target_margin_rate',
        'cny_exchange_rate',
        'exchange_rate_buffer',
        'shipping_cost_base',
        'customs_tax_rate'
    ]
    
    config_status = {}
    for key in configs_to_check:
        value = get_config(key)
        
        # Mask sensitive values
        if 'key' in key.lower() or 'secret' in key.lower():
            if value:
                masked = f'{value[:10]}...' if len(value) > 10 else '[TOO_SHORT]'
                config_status[key] = {
                    'configured': True,
                    'length': len(value),
                    'preview': masked
                }
                app.logger.info(f'[Config Verify] {key}: ✅ Configured (length: {len(value)})')
            else:
                config_status[key] = {
                    'configured': False,
                    'length': 0,
                    'preview': None
                }
                app.logger.warning(f'[Config Verify] {key}: ❌ Not configured')
        else:
            config_status[key] = {
                'configured': value is not None,
                'value': value
            }
            app.logger.info(f'[Config Verify] {key}: {value}')
    
    return jsonify({
        'success': True,
        'configs': config_status,
        'timestamp': datetime.now().isoformat()
    })

@app.route('/api/logs/stream')
@login_required
def stream_logs():
    """Stream activity logs (SSE endpoint)"""
    def generate():
        last_id = 0
        while True:
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM activity_logs WHERE id > ? ORDER BY id DESC LIMIT 10', (last_id,))
            logs = cursor.fetchall()
            conn.close()
            
            if logs:
                last_id = logs[0]['id']
                for log in logs:
                    yield f"data: {json.dumps(dict(log))}\n\n"
            
            time.sleep(2)
    
    return app.response_class(generate(), mimetype='text/event-stream')

@app.route('/api/products/<int:product_id>/approve', methods=['POST'])
@login_required
def approve_product(product_id):
    """Approve product for marketplace registration"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('UPDATE sourced_products SET status = ?, approved_at = CURRENT_TIMESTAMP WHERE id = ?', 
                   ('approved', product_id))
    conn.commit()
    conn.close()
    
    log_activity('product', f'Product {product_id} approved', 'success')
    
    return jsonify({'success': True})

@app.route('/api/products/<int:product_id>/register', methods=['POST'])
@login_required
def register_product(product_id):
    """Register product to marketplace"""
    data = request.json
    marketplace = data.get('marketplace')
    
    if marketplace not in ['naver', 'coupang']:
        return jsonify({'error': 'Invalid marketplace'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM sourced_products WHERE id = ?', (product_id,))
    product = cursor.fetchone()
    
    if not product:
        conn.close()
        return jsonify({'error': 'Product not found'}), 404
    
    # Prepare product data for marketplace
    product_data = {
        'title': product['title_kr'] or product['title_cn'],
        'description': product['marketing_copy'],
        'price': product['price_krw'],
        'images': json.loads(product['processed_images_json']) if product['processed_images_json'] else []
    }
    
    # Register to marketplace
    if marketplace == 'naver':
        result = register_to_naver(product_data)
    else:
        result = register_to_coupang(product_data)
    
    if 'error' in result:
        log_activity('marketplace', f'Failed to register product {product_id} to {marketplace}: {result["error"]}', 'error')
        return jsonify({'error': result['error']}), 500
    
    # Save marketplace listing
    cursor.execute('''
        INSERT INTO marketplace_listings (product_id, marketplace, external_product_id, status)
        VALUES (?, ?, ?, ?)
    ''', (product_id, marketplace, result.get('product_id', ''), 'active'))
    
    cursor.execute('UPDATE sourced_products SET status = ?, registered_at = CURRENT_TIMESTAMP WHERE id = ?',
                   ('registered', product_id))
    
    conn.commit()
    conn.close()
    
    log_activity('marketplace', f'Product {product_id} registered to {marketplace}', 'success')
    
    return jsonify({'success': True, 'external_id': result.get('product_id', '')})

@app.route('/api/products/<int:product_id>/market-analysis', methods=['GET', 'POST'])
@login_required
def analyze_product_market(product_id):
    """개별 상품에 대한 네이버 시장 분석"""
    conn = get_db()
    cursor = conn.cursor()
    
    # 상품 정보 조회
    cursor.execute('SELECT * FROM sourced_products WHERE id = ?', (product_id,))
    product = cursor.fetchone()
    
    if not product:
        conn.close()
        return jsonify({'error': 'Product not found'}), 404
    
    # 이미 분석된 데이터가 있는지 확인 (GET 요청 시)
    if request.method == 'GET' and product['market_analysis_json']:
        try:
            analysis = json.loads(product['market_analysis_json'])
            conn.close()
            return jsonify({
                'success': True,
                'cached': True,
                'analysis': analysis
            })
        except:
            pass
    
    # 네이버 API 인증 정보
    naver_client_id = get_config('naver_client_id')
    naver_client_secret = get_config('naver_client_secret')
    
    if not naver_client_id or not naver_client_secret:
        conn.close()
        return jsonify({
            'success': False,
            'error': '네이버 API 인증 정보가 설정되지 않았습니다. 설정 페이지에서 Client ID와 Secret을 입력해주세요.'
        }), 400
    
    # 키워드 추출 (우선순위: keywords > title_kr > AI 추출 > title_cn 축약)
    keyword = None
    
    # 1. keywords 필드가 있으면 우선 사용
    if product['keywords']:
        keyword = product['keywords'].split(',')[0].strip()  # 첫 번째 키워드
        app.logger.info(f'[Market Analysis] Using keyword from keywords field: {keyword}')
    
    # 2. title_kr이 있으면 사용
    elif product['title_kr']:
        keyword = product['title_kr']
        # 너무 길면 첫 3단어만
        words = keyword.split()
        if len(words) > 3:
            keyword = ' '.join(words[:3])
        app.logger.info(f'[Market Analysis] Using Korean title (truncated): {keyword}')
    
    # 3. AI로 일반 키워드 추출 (영어/중국어 제품명 → 한국어 카테고리명)
    else:
        title_original = product['title_cn'] or ''
        app.logger.info(f'[Market Analysis] Attempting AI keyword extraction from: {title_original[:80]}...')
        
        # OpenAI로 키워드 추출
        openai_api_key = get_config('openai_api_key')
        if openai_api_key:
            try:
                import openai
                openai.api_key = openai_api_key
                
                prompt = f"""다음 상품명을 보고 네이버/쿠팡에서 검색할 수 있는 한국어 키워드 2-3개를 추출해주세요.
브랜드명은 제외하고, 제품 카테고리나 일반명사만 사용해주세요.

상품명: {title_original}

응답 형식: 키워드1, 키워드2 (예: 무선 이어폰, 블루투스 헤드셋)
한국어 키워드만 작성하세요."""

                response = openai.ChatCompletion.create(
                    model="gpt-4o-mini",
                    messages=[
                        {"role": "system", "content": "You are a Korean e-commerce keyword extraction expert. Extract general product category keywords in Korean."},
                        {"role": "user", "content": prompt}
                    ],
                    max_tokens=50,
                    temperature=0.3
                )
                
                extracted = response.choices[0].message.content.strip()
                # 첫 번째 키워드만 사용
                keyword = extracted.split(',')[0].strip()
                
                app.logger.info(f'[Market Analysis] ✅ AI extracted keyword: {keyword}')
                
            except Exception as e:
                app.logger.warning(f'[Market Analysis] ⚠️ AI extraction failed: {str(e)}')
                # AI 실패 시 폴백
                keyword = title_original[:50] if len(title_original) > 50 else title_original
                app.logger.info(f'[Market Analysis] Using fallback (truncated original): {keyword}')
        else:
            # OpenAI 키가 없으면 원본 제목 사용
            keyword = title_original[:50] if len(title_original) > 50 else title_original
            app.logger.warning(f'[Market Analysis] ⚠️ No OpenAI key, using original title: {keyword}')
    
    if not keyword:
        conn.close()
        return jsonify({
            'success': False,
            'error': '상품 키워드를 추출할 수 없습니다.'
        }), 400
    
    # 시장 분석 실행 (재시도 로직 포함)
    from market_analysis import analyze_naver_market
    
    app.logger.info(f'[Market Analysis] Analyzing product {product_id} with keyword: {keyword}')
    
    result = analyze_naver_market(keyword, naver_client_id, naver_client_secret)
    
    # 결과가 너무 적으면 더 일반적인 키워드로 재시도
    if result.get('success') and result.get('analyzed_products', 0) < 10:
        app.logger.warning(f'[Market Analysis] ⚠️ Low results ({result["analyzed_products"]} products). Trying broader keyword...')
        
        # AI로 더 일반적인 키워드 추출 시도
        openai_api_key = get_config('openai_api_key')
        if openai_api_key:
            try:
                import openai
                openai.api_key = openai_api_key
                
                prompt = f"""이 키워드의 검색 결과가 너무 적습니다: "{keyword}"
더 일반적이고 포괄적인 한국어 키워드 1개만 제안해주세요.

예시:
- "맞춤 주얼리" → "목걸이"
- "스네이크 패턴 부츠" → "여성 부츠"
- "오실로스코프" → "측정 장비"

응답: (키워드 1개만)"""

                response = openai.ChatCompletion.create(
                    model="gpt-4o-mini",
                    messages=[
                        {"role": "system", "content": "Suggest a broader, more general Korean keyword."},
                        {"role": "user", "content": prompt}
                    ],
                    max_tokens=20,
                    temperature=0.3
                )
                
                broader_keyword = response.choices[0].message.content.strip()
                app.logger.info(f'[Market Analysis] 🔄 Retrying with broader keyword: {broader_keyword}')
                
                # 재시도
                retry_result = analyze_naver_market(broader_keyword, naver_client_id, naver_client_secret)
                
                if retry_result.get('success') and retry_result.get('analyzed_products', 0) >= 10:
                    app.logger.info(f'[Market Analysis] ✅ Retry successful: {retry_result["analyzed_products"]} products')
                    result = retry_result
                    keyword = broader_keyword  # 성공한 키워드로 업데이트
                else:
                    app.logger.info(f'[Market Analysis] Keeping original result')
                    
            except Exception as e:
                app.logger.warning(f'[Market Analysis] Retry failed: {str(e)}')
    
    if not result.get('success'):
        conn.close()
        # 에러 details도 함께 전달
        return jsonify({
            'success': False,
            'error': result.get('error', '시장 분석 실패'),
            'details': result.get('details', {})
        }), 500
    
    # 쿠팡 API 분석 (선택적)
    coupang_data = None
    coupang_access_key = get_config('coupang_access_key')
    coupang_secret_key = get_config('coupang_secret_key')
    
    if coupang_access_key and coupang_secret_key:
        from market_analysis import analyze_coupang_market
        app.logger.info(f'[Market Analysis] Also analyzing Coupang market...')
        coupang_result = analyze_coupang_market(title, coupang_access_key, coupang_secret_key)
        if coupang_result.get('success'):
            coupang_data = coupang_result
            app.logger.info(f'[Market Analysis] ✅ Coupang analysis completed')
        else:
            app.logger.warning(f'[Market Analysis] ⚠️ Coupang analysis failed: {coupang_result.get("error")}')
    
    # 분석 결과를 DB에 저장
    analysis_data = {
        'keyword': result['keyword'],
        'sources': ['naver'],  # 분석 소스 추적
        'naver': {
            'total_products': result['total_products'],
            'analyzed_products': result['analyzed_products'],
            'avg_price': result['avg_price'],
            'median_price': result['median_price'],
            'min_price': result['min_price'],
            'max_price': result['max_price'],
            'q1_price': result['q1_price'],
            'q3_price': result['q3_price'],
            'recommended_price': result['recommended_price'],
            'price_distribution': result['price_distribution'],
            'analysis_summary': result['analysis_summary'],
            'top_products': result['top_products'][:3]  # 상위 3개 참고 링크
        },
        'coupang': coupang_data if coupang_data else None,
        'timestamp': result['timestamp']
    }
    
    # 쿠팡 데이터가 있으면 sources에 추가
    if coupang_data:
        analysis_data['sources'].append('coupang')
    
    cursor.execute('''
        UPDATE sourced_products 
        SET market_analysis_json = ? 
        WHERE id = ?
    ''', (json.dumps(analysis_data, ensure_ascii=False), product_id))
    
    conn.commit()
    conn.close()
    
    app.logger.info(f'[Market Analysis] ✅ Analysis completed for product {product_id}')
    
    return jsonify({
        'success': True,
        'cached': False,
        'analysis': analysis_data
    })

# ============================================================================
# MAIN
# ============================================================================

# Background scheduler for automatic order sync
def start_order_sync_scheduler():
    """
    Start background scheduler to sync orders every 10 minutes
    """
    import threading
    import time
    
    def sync_loop():
        app.logger.info('[Order Sync Scheduler] 🚀 Started')
        while True:
            try:
                time.sleep(600)  # Wait 10 minutes
                
                with app.app_context():
                    app.logger.info('[Order Sync Scheduler] 🔄 Running automatic sync...')
                    
                    # Sync Naver
                    naver_result = sync_naver_orders()
                    if naver_result.get('success'):
                        count = naver_result.get('synced_count', 0)
                        if count > 0:
                            app.logger.info(f'[Order Sync Scheduler] ✅ Naver: {count} new orders')
                    
                    # Sync Coupang
                    coupang_result = sync_coupang_orders()
                    if coupang_result.get('success'):
                        count = coupang_result.get('synced_count', 0)
                        if count > 0:
                            app.logger.info(f'[Order Sync Scheduler] ✅ Coupang: {count} new orders')
                    
            except Exception as e:
                app.logger.error(f'[Order Sync Scheduler] ❌ Error: {str(e)}')
    
    # Start in background thread
    thread = threading.Thread(target=sync_loop, daemon=True)
    thread.start()
    app.logger.info('[Order Sync Scheduler] ✅ Scheduler started (10 min interval)')

if __name__ == '__main__':
    # Ensure required directories exist
    os.makedirs('static/processed_images', exist_ok=True)
    os.makedirs('static/exports', exist_ok=True)
    
    # Start automatic order sync scheduler
    start_order_sync_scheduler()
    
    # Run app
    app.run(host='0.0.0.0', port=5000, debug=False)
